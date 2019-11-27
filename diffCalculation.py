import openpyxl
#from datetime import datetime
import pandas as pd
import numpy as np
from copy import copy
import datetime
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW
import string
import time
from openpyxl.utils import FORMULAE
from openpyxl.styles import NamedStyle
import random



def loadBook():
  global book
  book = openpyxl.load_workbook(r'C:\Users\shoshana\PycharmProjects\pandas\SeptemberAfterMonitoringBeforeManipilation.xlsx')

  return

def creatingNewWorkBook():
  if x ==earliest:
       global fileAfterMonitoring2
       fileAfterMonitoring2 = openpyxl.Workbook()

  fileAfterMonitoring2.create_sheet(x)

  return

def variablesPrepraion():

  global now
  now = datetime.datetime.now()
  global today
  today = datetime.date.today()
  global d1
  d1 = today.strftime("%d/%m/%Y")

  return


def sheetsLoadingTyping(book, trainNameColumn, exitTimeColumn, arrivalTimeColumn):

   book.active
   global realTime_1
   realTime_1 = [None] * 350
   global exitTime_1
   exitTime_1 = [None] * 350
   global sheetsSouth
   sheetsSouth = []
   global realTimeSouth
   realTimeSouth = [None] * 350
   global exitTimeSouth
   exitTimeSouth = [None] * 350
   global BlueExitTimeARR
   BlueExitTimeARR = [None] * 350
   global trainNameBlue
   trainNameBlue = [None] * 350
   global trainNameNonBlue
   trainNameNonBlue = [None] * 350
   global arrivalTimeBlue
   arrivalTimeBlue = [None] * 350
   global arrivalTimeNonBlue
   arrivalTimeNonBlue = [None] * 350
   global aftermidnight
   aftermidnight = []
   global trainNameNonBlueaftermidnight
   trainNameNonBlueaftermidnight = []
   global arrivalTimeNonBlueaftermidnight
   arrivalTimeNonBlueaftermidnight = []

   global CounterHowMuchBlueIsThere
   CounterHowMuchBlueIsThere = 0




   for l in range(0, 250): #go over all cells

        if (not(book[x].cell(row=l + 5, column=exitTimeColumn).font.color.rgb =="FF00B0F0")): #check if color is blue
          if (book[x].cell(row=l + 5, column=exitTimeColumn).value) is not None:#check that cell is not None
              #clean from cells needless date as "30/12/1899 or 01/01/1900
              #then, contain in arrays all 3 column of typing report
              if (type(book[x].cell(row=l + 5, column=exitTimeColumn).value)) is datetime.time:
                  dateTimeAsStr = datetime.time.strftime((book[x].cell(row=l + 5, column=exitTimeColumn).value), "%d/%m/%Y %H:%M:%S")

                  if "00:" in dateTimeAsStr or "01:" in dateTimeAsStr or "02:" in dateTimeAsStr:
                          for i in range(0, 6):
                              i = str(i)
                              for j in range(0, 10):
                                  j = str(j)
                                  for k in range(0, 6):
                                      k = str(k)
                                      for m in range(0, 10):
                                          m = str(m)

                                          if ("00:" + i + j + ":" + k + m) in dateTimeAsStr or (
                                                  "01:" + i + j + ":" + k + m) in dateTimeAsStr or (
                                                  "02:" + i + j + ":" + k + m) in dateTimeAsStr:
                                              if ("1899/12/30" in dateTimeAsStr):
                                                  dateTimeAsStr = dateTimeAsStr.replace("1899/12/30 ", "")
                                              if ("01/01/1900" in dateTimeAsStr):
                                                  dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")

                                              aftermidnight.append(dateTimeAsStr)
                                              trainNameNonBlueaftermidnight.append((book[x].cell(row=l + 5, column=trainNameColumn).value))
                                              arrivalTimeNonBlueaftermidnight.append((book[x].cell(row=l + 5, column=arrivalTimeColumn).value))

                                          else:
                                              exitTime_1[l] = (book[x].cell(row=l + 5, column=exitTimeColumn).value)
                                              trainNameNonBlue[l] = (
                                                  book[x].cell(row=l + 5, column=trainNameColumn).value)
                                              arrivalTimeNonBlue[l] = (
                                                  book[x].cell(row=l + 5, column=arrivalTimeColumn).value)


                  else:
                    if ("1899/12/30" in dateTimeAsStr):
                        dateTimeAsStr =dateTimeAsStr.replace("1899/12/30 ", "")
                    if ("01/01/1900" in dateTimeAsStr):
                        dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")
                    exitTime_1[l]=(dateTimeAsStr)
                    trainNameNonBlue[l] = (book[x].cell(row=l + 5, column=trainNameColumn).value)
                    arrivalTimeNonBlue[l] = (book[x].cell(row=l + 5, column=arrivalTimeColumn).value)

              elif  (type(book[x].cell(row=l + 5, column=exitTimeColumn).value)) is str:
                  if "00:" in (book[x].cell(row=l + 5, column=exitTimeColumn).value) or "01:" in (
                  book[x].cell(row=l + 5, column=exitTimeColumn).value) or "02:" in (
                  book[x].cell(row=l + 5, column=exitTimeColumn).value):
                    dateTimeAsStr = book[x].cell(row=l + 5, column=exitTimeColumn).value

                    for i in range(0, 6):
                        i = str(i)
                        for j in range(0, 10):
                            j = str(j)
                            for k in range(0, 6):
                                k = str(k)
                                for m in range(0, 10):
                                    m = str(m)

                        if ("00:" + i + j + ":" + k + m) in dateTimeAsStr or ("01:" + i + j + ":" + k + m) in dateTimeAsStr or ("02:" + i + j + ":" + k + m) in dateTimeAsStr:
                            if ("1899/12/30" in dateTimeAsStr):
                                dateTimeAsStr = dateTimeAsStr.replace("1899/12/30 ", "")
                            if ("01/01/1900" in dateTimeAsStr):
                                dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")

                            aftermidnight.append(dateTimeAsStr)
                            trainNameNonBlueaftermidnight.append((book[x].cell(row=l + 5, column=trainNameColumn).value))
                            arrivalTimeNonBlueaftermidnight.append((book[x].cell(row=l + 5, column=arrivalTimeColumn).value))

                        else:
                            exitTime_1[l] = (book[x].cell(row=l + 5, column=exitTimeColumn).value)
                            trainNameNonBlue[l] = (book[x].cell(row=l + 5, column=trainNameColumn).value)
                            arrivalTimeNonBlue[l] = (book[x].cell(row=l + 5, column=arrivalTimeColumn).value)
                  else:
                      exitTime_1[l] = (book[x].cell(row=l + 5, column=exitTimeColumn).value)
                      trainNameNonBlue[l] = (book[x].cell(row=l + 5, column=trainNameColumn).value)
                      arrivalTimeNonBlue[l] = (book[x].cell(row=l + 5, column=arrivalTimeColumn).value)

              elif "00:" in datetime.datetime.strftime((book[x].cell(row=l + 5, column=exitTimeColumn).value), "%d/%m/%Y %H:%M:%S") or "01:" in datetime.datetime.strftime((book[x].cell(row=l + 5, column=exitTimeColumn).value), "%d/%m/%Y %H:%M:%S") or "02:" in datetime.datetime.strftime((book[x].cell(row=l + 5, column=exitTimeColumn).value), "%d/%m/%Y %H:%M:%S"):
                  dateTimeAsStr = datetime.datetime.strftime((book[x].cell(row=l + 5, column=exitTimeColumn).value), "%d/%m/%Y %H:%M:%S")

                  for i in range(0, 6):
                      i = str(i)
                      for j in range(0, 10):
                          j = str(j)
                          for k in range(0, 6):
                              k = str(k)
                              for m in range(0, 10):
                                  m = str(m)

                      if ("00:" + i + j + ":" + k + m) in dateTimeAsStr or ("01:" + i + j + ":" + k + m) in dateTimeAsStr or ("02:" + i + j + ":" + k + m) in dateTimeAsStr:
                          if ("1899/12/30" in dateTimeAsStr):
                              dateTimeAsStr = dateTimeAsStr.replace("1899/12/30 ", "")
                          if ("01/01/1900" in dateTimeAsStr):
                              dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")

                          aftermidnight.append(dateTimeAsStr)
                          trainNameNonBlueaftermidnight.append((book[x].cell(row=l + 5, column=trainNameColumn).value))
                          arrivalTimeNonBlueaftermidnight.append(
                              (book[x].cell(row=l + 5, column=arrivalTimeColumn).value))

                      else:
                          exitTime_1[l] = (book[x].cell(row=l + 5, column=exitTimeColumn).value)
                          trainNameNonBlue[l] = (book[x].cell(row=l + 5, column=trainNameColumn).value)
                          arrivalTimeNonBlue[l] = (book[x].cell(row=l + 5, column=arrivalTimeColumn).value)



              else:
               exitTime_1[l]=(book[x].cell(row=l + 5, column=exitTimeColumn).value)
               trainNameNonBlue[l]=(book[x].cell(row=l + 5, column=trainNameColumn).value)
               arrivalTimeNonBlue[l]=(book[x].cell(row=l + 5, column=arrivalTimeColumn).value)



      # if it's indeed blue contain it in another arrays

        elif (book[x].cell(row=l + 5, column=exitTimeColumn).font.color.rgb =="FF00B0F0"):

          if (book[x].cell(row=l + 5, column=exitTimeColumn).value) is not None:

              if (type(book[x].cell(row=l + 5, column=exitTimeColumn).value)) is datetime.time:
                  dateTimeAsStr = datetime.time.strftime((book[x].cell(row=l + 5, column=exitTimeColumn).value),
                                                         "%d/%m/%Y %H:%M:%S")
                  if ("1899/12/30" in dateTimeAsStr):
                      dateTimeAsStr = dateTimeAsStr.replace("1899/12/30 ", "")
                  if ("01/01/1900" in dateTimeAsStr):
                      dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")

                  BlueExitTimeARR[l]=(dateTimeAsStr)
              else:
                  BlueExitTimeARR[l]=(book[x].cell(row=l + 5, column=exitTimeColumn).value)
              trainNameBlue[l]=(book[x].cell(row=l + 5, column=trainNameColumn).value)
              arrivalTimeBlue[l]=(book[x].cell(row=l + 5, column=arrivalTimeColumn).value)
              CounterHowMuchBlueIsThere = CounterHowMuchBlueIsThere +1

        else:
          if (book[x].cell(row=l + 5, column=exitTimeColumn).value) is not None:#check that cell is not None
              #clean from cells needless date as "30/12/1899 or 01/01/1900
              #then, contain in arrays all 3 column of typing report
              if (type(book[x].cell(row=l + 5, column=exitTimeColumn).value)) is datetime.time:
                  dateTimeAsStr = datetime.time.strftime((book[x].cell(row=l + 5, column=exitTimeColumn).value), "%d/%m/%Y %H:%M:%S")
                  if ("1899/12/30" in dateTimeAsStr):
                      dateTimeAsStr =dateTimeAsStr.replace("1899/12/30 ", "")
                  if ("01/01/1900" in dateTimeAsStr):
                      dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")
                  exitTime_1[l]=(dateTimeAsStr)

              else:
                  exitTime_1[l]=(book[x].cell(row=l + 5, column=exitTimeColumn).value)
              trainNameNonBlue[l]=(book[x].cell(row=l + 5, column=trainNameColumn).value)
              arrivalTimeNonBlue[l]=(book[x].cell(row=l + 5, column=arrivalTimeColumn).value)


   trainNameNonBlue = firstDeletingOfNones(trainNameNonBlue)
   arrivalTimeNonBlue = firstDeletingOfNones(arrivalTimeNonBlue)
   decideIfNeddSotringOrNot = 0

   for y in range(0 , len(aftermidnight)):
       if aftermidnight[y] is not None and aftermidnight[y] != 'None':
           decideIfNeddSotringOrNot = decideIfNeddSotringOrNot + 3
       else:
           decideIfNeddSotringOrNot = decideIfNeddSotringOrNot + 0


   try:
       exitTime_1, trainNameNonBlue = sortingArraysByMainArray(exitTime_1, trainNameNonBlue, doYouWantToSortMainArr=1)
   except Exception:
       pass

   try:
       exitTime_1, arrivalTimeNonBlue = sortingArraysByMainArray(exitTime_1, arrivalTimeNonBlue, doYouWantToSortMainArr=0)
   except Exception:
       pass



   if decideIfNeddSotringOrNot > 0:
       aftermidnight = firstDeletingOfNones(aftermidnight)
       trainNameNonBlueaftermidnight = firstDeletingOfNones(trainNameNonBlueaftermidnight)
       arrivalTimeNonBlueaftermidnight = firstDeletingOfNones(arrivalTimeNonBlueaftermidnight)
       exitTime_1 = exitTime_1 + aftermidnight
       trainNameNonBlue = trainNameNonBlue + trainNameNonBlueaftermidnight
       arrivalTimeNonBlue = arrivalTimeNonBlue + arrivalTimeNonBlueaftermidnight


   return

def firstDeletingOfNones(Arr):

   for i in range(0, len(Arr)):

       if Arr[i] is not None and Arr[i] != "None":

           theBiggestINdexThatIsNotEmpty = i

   Arr = Arr[0: theBiggestINdexThatIsNotEmpty + 1]

   return(Arr)

def sortingArraysByMainArray(mainArr, secondaryArray, doYouWantToSortMainArr):
   twoArrssorted = sorted(zip(mainArr, secondaryArray))
   secondaryArray = [secondaryArray for mainArr, secondaryArray in twoArrssorted]
   if doYouWantToSortMainArr == 0:
       mainArr = sorted(mainArr)
       #mainArr = mainArr + ArrOfLastValueOfMainArr
       #secondaryArray = secondaryArray + ArrOfLastValueOfSecondaryArr



   return(mainArr ,secondaryArray)


def sheetsLoadingfranchisee(book, dateColumn, RBColumn, stopPointColumn, theoTimeColumn, applTimeColumn, realTimeColumn):

   global RedRealTimeARR
   RedRealTimeARR = [None] * 350
   global Date
   Date = [None] * 350
   global redDate
   redDate = [None] * 350
   global RB
   RB = [None] * 350
   global redRB
   redRB = [None] * 350
   global theoTime
   theoTime = [None] * 350
   global redTheoTime
   redTheoTime = [None] * 350
   global apllTime
   apllTime = [None] * 350
   global redApllTime
   redApllTime = [None] * 350
   global stopPoint
   stopPoint = [None] * 350
   global redStopPoint
   redStopPoint = [None] * 350
   global realTime_1
   realTime_1 = [None] *350
   global aftermidnightRealTime
   aftermidnightRealTime = []
   global Dateaftermidnight
   Dateaftermidnight = []
   global RBaftermidnight
   RBaftermidnight = []
   global theoTimeaftermidnight
   theoTimeaftermidnight = []
   global apllTimeftermidnight
   apllTimeftermidnight = []
   global stopPointaftermidnight
   stopPointaftermidnight = []
   global CounterHowMuchRedIsThere
   CounterHowMuchRedIsThere = 0



   for l in range(0, 250): #go over all cells

       if (not (book[x].cell(row=l + 5, column=stopPointColumn).font.color.rgb == "FFFF0000")): #check if cell color is not red
           if (book[x].cell(row=l + 5, column=stopPointColumn).value) is not None:


               # clean from cells needless date as "30/12/1899" or "01/01/1900"
               # then, contain in arrays all 6 column of franchisee report

               if (type(book[x].cell(row=l + 5, column=realTimeColumn).value)) is datetime.time:
                   dateTimeAsStr = datetime.time.strftime((book[x].cell(row=l + 5, column=realTimeColumn).value),"%d/%m/%Y %H:%M:%S")
                   if "00:" in dateTimeAsStr or "01:" in dateTimeAsStr or "02:" in dateTimeAsStr:
                       for i in range(0, 6):
                           i = str(i)
                           for j in range(0, 10):
                               j = str(j)
                               for k in range(0, 6):
                                   k = str(k)
                                   for m in range(0, 10):
                                       m = str(m)

                                       if ("00:" + i + j + ":" + k + m) in dateTimeAsStr or ("01:" + i + j + ":" + k + m) in dateTimeAsStr or ("02:" + i + j + ":" + k + m) in dateTimeAsStr:

                                           if ("1899/12/30" in dateTimeAsStr):
                                               dateTimeAsStr = dateTimeAsStr.replace("1899/12/30 ", "")
                                           if ("01/01/1900" in dateTimeAsStr):
                                               dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")

                                           aftermidnightRealTime.append(dateTimeAsStr)
                                           Dateaftermidnight.append(book[x].cell(row=l + 5, column=dateColumn).value)
                                           RBaftermidnight.append(book[x].cell(row=l + 5, column=RBColumn).value)
                                           theoTimeaftermidnight.append(book[x].cell(row=l + 5, column=theoTimeColumn).value)
                                           apllTimeftermidnight.append(book[x].cell(row=l + 5, column=applTimeColumn).value)
                                           stopPointaftermidnight.append(book[x].cell(row=l + 5, column=stopPointColumn).value)

                                       else:
                                           realTime_1[l] = (book[x].cell(row=l + 5, column=realTimeColumn).value)
                                           Date[l] = (book[x].cell(row=l + 5, column=dateColumn).value)
                                           RB[l] = (book[x].cell(row=l + 5, column=RBColumn).value)
                                           theoTime[l] = (book[x].cell(row=l + 5, column=theoTimeColumn).value)
                                           apllTime[l] = (book[x].cell(row=l + 5, column=applTimeColumn).value)
                                           stopPoint[l] = (book[x].cell(row=l + 5, column=stopPointColumn).value)




                   else:
                       if ("1899/12/30" in dateTimeAsStr):
                            dateTimeAsStr = dateTimeAsStr.replace("1899/12/30 ", "")
                       if ("01/01/1900" in dateTimeAsStr):
                            dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")

                       realTime_1[l]=(dateTimeAsStr)
                       Date[l] = (book[x].cell(row=l + 5, column=dateColumn).value)
                       RB[l] = (book[x].cell(row=l + 5, column=RBColumn).value)
                       theoTime[l] = (book[x].cell(row=l + 5, column=theoTimeColumn).value)
                       apllTime[l] = (book[x].cell(row=l + 5, column=applTimeColumn).value)
                       stopPoint[l] = (book[x].cell(row=l + 5, column=stopPointColumn).value)

               elif "00:" in book[x].cell(row=l + 5, column=realTimeColumn).value or "01:" in book[x].cell(row=l + 5, column=realTimeColumn).value or "02:" in book[x].cell(row=l + 5, column=realTimeColumn).value:
                   dateTimeAsStr = book[x].cell(row=l + 5, column=realTimeColumn).value
                   for i in range(0, 6):
                    i = str(i)
                    for j in range(0, 10):
                        j = str(j)
                        for k in range(0, 6):
                            k = str(k)
                            for m in range(0, 10):
                                m = str(m)

                                if ("00:" + i + j + ":" + k + m) in dateTimeAsStr or ("01:" + i + j + ":" + k + m) in dateTimeAsStr or ("02:" + i + j + ":" + k + m) in dateTimeAsStr:

                                    if ("1899/12/30" in dateTimeAsStr):
                                        dateTimeAsStr = dateTimeAsStr.replace("1899/12/30 ", "")
                                    if ("01/01/1900" in dateTimeAsStr):
                                        dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")

                                    aftermidnightRealTime.append(dateTimeAsStr)
                                    Dateaftermidnight.append(book[x].cell(row=l + 5, column=dateColumn).value)
                                    RBaftermidnight.append(book[x].cell(row=l + 5, column=RBColumn).value)
                                    theoTimeaftermidnight.append(book[x].cell(row=l + 5, column=theoTimeColumn).value)
                                    apllTimeftermidnight.append(book[x].cell(row=l + 5, column=applTimeColumn).value)
                                    stopPointaftermidnight.append(book[x].cell(row=l + 5, column=stopPointColumn).value)

                                else:
                                    realTime_1[l] = (book[x].cell(row=l + 5, column=realTimeColumn).value)
                                    Date[l] = (book[x].cell(row=l + 5, column=dateColumn).value)
                                    RB[l] = (book[x].cell(row=l + 5, column=RBColumn).value)
                                    theoTime[l] = (book[x].cell(row=l + 5, column=theoTimeColumn).value)
                                    apllTime[l] = (book[x].cell(row=l + 5, column=applTimeColumn).value)
                                    stopPoint[l] = (book[x].cell(row=l + 5, column=stopPointColumn).value)

               else:
                   if (book[x].cell(row=l + 5, column=realTimeColumn).value) is not None and (book[x].cell(row=l + 5, column=dateColumn).value) is not None:
                       realTime_1[l]=(book[x].cell(row=l + 5, column=realTimeColumn).value)
                       Date[l]=(book[x].cell(row=l + 5, column=dateColumn).value)
                       RB[l]=(book[x].cell(row=l + 5, column=RBColumn).value)
                       theoTime[l]=(book[x].cell(row=l + 5, column=theoTimeColumn).value)
                       apllTime[l]=(book[x].cell(row=l + 5, column=applTimeColumn).value)
                       stopPoint[l]=(book[x].cell(row=l + 5, column=stopPointColumn).value)


       elif(book[x].cell(row=l + 5, column=stopPointColumn).value) != "S23HERZ1":


           if (book[x].cell(row=l + 5, column=realTimeColumn).value) is not None and (book[x].cell(row=l + 5, column=dateColumn).value) is not None:
               if (type(book[x].cell(row=l + 5, column=realTimeColumn).value)) is datetime.time:
                   dateTimeAsStr = datetime.time.strftime((book[x].cell(row=l + 5, column=realTimeColumn).value),
                                                          "%d/%m/%Y %H:%M:%S")
                   if ("1899/12/30" in dateTimeAsStr):
                       dateTimeAsStr = dateTimeAsStr.replace("1899/12/30 ", "")
                   if ("01/01/1900" in dateTimeAsStr):
                       dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")

                   RedRealTimeARR[l]=(dateTimeAsStr)
               else:
                   RedRealTimeARR[l]=(book[x].cell(row=l + 5, column=realTimeColumn).value)

               redDate[l]=(book[x].cell(row=l + 5, column=dateColumn).value)
               redRB[l]=(book[x].cell(row=l + 5, column=RBColumn).value)
               redTheoTime[l]=(book[x].cell(row=l + 5, column=theoTimeColumn).value)
               redApllTime[l]=(book[x].cell(row=l + 5, column=applTimeColumn).value)
               redStopPoint[l]=(book[x].cell(row=l + 5, column=stopPointColumn).value)
               CounterHowMuchRedIsThere = CounterHowMuchRedIsThere +1
       else:

          if (book[x].cell(row=l + 5, column=stopPointColumn).value) is not None:

               # clean from cells needless date as "30/12/1899" or "01/01/1900"
               # then, contain in arrays all 6 column of franchisee report

               if (type(book[x].cell(row=l + 5, column=realTimeColumn).value)) is datetime.time:
                   dateTimeAsStr = datetime.time.strftime((book[x].cell(row=l + 5, column=realTimeColumn).value),
                                                          "%d/%m/%Y %H:%M:%S")

                   if ("1899/12/30" in dateTimeAsStr):
                       dateTimeAsStr = dateTimeAsStr.replace("1899/12/30 ", "")
                   if ("01/01/1900" in dateTimeAsStr):
                       dateTimeAsStr = dateTimeAsStr.replace("01/01/1900 ", "")

                   realTime_1[l]=(dateTimeAsStr)

               elif (book[x].cell(row=l + 5, column=stopPointColumn).value) is not None:
                   realTime_1[l]=(book[x].cell(row=l + 5, column=realTimeColumn).value)
                   Date[l]=(book[x].cell(row=l + 5, column=dateColumn).value)
                   RB[l]=(book[x].cell(row=l + 5, column=RBColumn).value)
                   theoTime[l]=(book[x].cell(row=l + 5, column=theoTimeColumn).value)
                   apllTime[l]=(book[x].cell(row=l + 5, column=applTimeColumn).value)
                   stopPoint[l]=(book[x].cell(row=l + 5, column=stopPointColumn).value)

   global ArrCounterHowMuchRedIsThere
   ArrCounterHowMuchRedIsThere = []

   if CounterHowMuchRedIsThere > 0:
       for r in range(0, CounterHowMuchBlueIsThere):
           ArrCounterHowMuchRedIsThere.append(" ")
   decideIfNeddSotringOrNot = 0
   for y in range(0, len(aftermidnightRealTime)):
       if aftermidnightRealTime[y] is not None and aftermidnightRealTime[y] != 'None':
           decideIfNeddSotringOrNot = decideIfNeddSotringOrNot + 3
       else:
           decideIfNeddSotringOrNot = decideIfNeddSotringOrNot + 0

   Date = firstDeletingOfNones(Date)
   Date = [i for i in Date if i]
   realTime_1 = firstDeletingOfNones(realTime_1)
   realTime_1 = [i for i in realTime_1 if i]

   realTime_1, Date = sortingArraysByMainArray(realTime_1, Date, doYouWantToSortMainArr=1)

   RB = firstDeletingOfNones(RB)
   RB = [i for i in RB if i]

   realTime_1, RB = sortingArraysByMainArray(realTime_1, RB, doYouWantToSortMainArr=1)

   stopPoint = firstDeletingOfNones(stopPoint)
   stopPoint = [i for i in stopPoint if i]

   realTime_1, stopPoint = sortingArraysByMainArray(realTime_1, stopPoint, doYouWantToSortMainArr=1)

   theoTime = firstDeletingOfNones(theoTime)
   theoTime = [i for i in theoTime if i]

   realTime_1, theoTime = sortingArraysByMainArray(realTime_1, theoTime, doYouWantToSortMainArr=1)

   apllTime = firstDeletingOfNones(apllTime)
   apllTime = [i for i in apllTime if i]

   realTime_1, apllTime = sortingArraysByMainArray(realTime_1, apllTime, doYouWantToSortMainArr=0)

   ArrCounterHowMuchBlueIsThere=[]
   if CounterHowMuchBlueIsThere > 0:
      for r in range(0,CounterHowMuchBlueIsThere):
        ArrCounterHowMuchBlueIsThere.append(" ")
      realTime_1 = realTime_1 + ArrCounterHowMuchBlueIsThere
      Date = Date + ArrCounterHowMuchBlueIsThere
      RB = RB + ArrCounterHowMuchBlueIsThere
      theoTime = theoTime + ArrCounterHowMuchBlueIsThere
      apllTime = apllTime + ArrCounterHowMuchBlueIsThere
      stopPoint = stopPoint + ArrCounterHowMuchBlueIsThere


   if decideIfNeddSotringOrNot > 0:
       aftermidnightRealTime = firstDeletingOfNones(aftermidnightRealTime)
       Dateaftermidnight = firstDeletingOfNones(Dateaftermidnight)
       RBaftermidnight = firstDeletingOfNones(RBaftermidnight)
       theoTimeaftermidnight = firstDeletingOfNones(theoTimeaftermidnight)
       apllTimeftermidnight = firstDeletingOfNones(apllTimeftermidnight)
       stopPointaftermidnight = firstDeletingOfNones(stopPointaftermidnight)


       realTime_1 = realTime_1 + aftermidnightRealTime
       Date = Date + Dateaftermidnight
       RB = RB + RBaftermidnight
       theoTime = theoTime + theoTimeaftermidnight
       apllTime = apllTime + apllTimeftermidnight
       stopPoint = stopPoint + stopPointaftermidnight

   if realTime_1[0] == realTime_1[len(realTime_1)-1]:
       realTime_1= realTime_1[1: len(realTime_1)]
       Date = Date[1:len(Date)]
       RB = RB[1:len(RB)]
       theoTime = theoTime[1:len(theoTime)]
       apllTime = apllTime[1:len(apllTime)]
       stopPoint = stopPoint[1:len(stopPoint)]



   #prepration for next functions:
   if realTimeColumn ==7:
       global exitTimeSouth
       exitTimeSouth = exitTime_1
       global realTimeSouth
       realTimeSouth =realTime_1
       global minLen
       minLen = min(len(realTimeSouth), len(exitTimeSouth))
   else:
       global exitTimeNorth
       exitTimeNorth = exitTime_1
       global realTimeNorth
       realTimeNorth = realTime_1
       minLen = min(len(realTimeNorth), len(exitTimeNorth))

   return

def calculationDiffTime(realTime_1, exitTime_1, realTimeDirection, exitTimeDirection, tdeltaColumn, arrivalTimeNonBlue,  trainNameNonBlue, Date, RB, stopPoint, theoTime, apllTime, south):
   realTime = realTime_1
   exitTime = exitTime_1
   #in order to calculate diff time, I'm adding date of today
   try:
       for l in range(0, len(realTime)):
           realTime[l] = datetime.datetime.strptime(d1 + " " + realTime[l], "%d/%m/%Y %H:%M:%S")
   except Exception:
       pass
   try:
       for l in range(0, len(exitTime)):
           exitTime[l] = datetime.datetime.strptime(d1 + " " + exitTime[l], "%d/%m/%Y %H:%M:%S")
   except Exception:
       pass

   emptyArr = [None] * 250
   realTime = realTime + emptyArr
   exitTime = exitTime + emptyArr
   m = 0
   c = 0
   d = 0
   g = 0

   global DateARR
   DateARR = [None] * 350
   global RBARR
   RBARR = [None] * 350
   global stopPointARR
   stopPointARR = [None] * 350
   global theoTimeARR
   theoTimeARR = [None] * 350
   global apllTimeARR
   apllTimeARR = [None] * 350
   global realTimeARR
   realTimeARR = [None] * 350
   global trainNameARR
   trainNameARR = [None] * 350
   global exitTimeARR
   exitTimeARR = [None] * 350
   global arrivalTimeARR
   arrivalTimeARR = [None] * 350
   status = "unknown"
   global tdeltaARR
   tdeltaARR = [None] * 350
   whatToMarkInYellowBackground =[]

   arrivalTimeNonBlue = arrivalTimeNonBlue + emptyArr
   trainNameNonBlue =  trainNameNonBlue + emptyArr
   Date = Date + emptyArr
   RB = RB + emptyArr
   stopPoint = stopPoint + emptyArr
   theoTime = theoTime + emptyArr
   apllTime = apllTime + emptyArr

   exitTime = exitTime + emptyArr
   realTime = realTime + emptyArr

   for t in range(0, 250):
       #claculate diff time

       if (g < len(realTimeDirection)) and (m < len(exitTimeDirection) and (realTimeDirection[g]) is not None and (exitTimeDirection[m]) is not None) and realTimeDirection[g] != " ":
           tdelta = ((max(realTimeDirection[g], exitTimeDirection[m])) - (min(realTimeDirection[g], exitTimeDirection[m])))


       exitTimeInMinuets = (exitTime[m])
       RealTimeInMinuets = (realTime[g])

       if t!= 250:
           nextExitTimeInMinuets = (exitTime[m+1])
           nextRealTimeInMinuets = (realTime[g+1])

       b = datetime.timedelta(hours=2) / 60


       # check if diff time is more than 2 minuets
       statusOfCorrection = 0
       if ((exitTimeInMinuets is not None) and (RealTimeInMinuets is not None)) or ((exitTimeInMinuets != "None")) or ((RealTimeInMinuets != "None")) and RealTimeInMinuets != " ":
         if RealTimeInMinuets is not None and RealTimeInMinuets != " ":
               if exitTimeInMinuets is not None:

                   if (((exitTimeInMinuets - (b)) > RealTimeInMinuets) or ((exitTimeInMinuets + (b)) < RealTimeInMinuets)) and (t > 0):

                       previousExitTimeInMinuets = exitTime[m-1]
                       prviousRealTimeInMinuets = realTime[g-1]

                       if (((previousExitTimeInMinuets - (b)) > prviousRealTimeInMinuets) or ((previousExitTimeInMinuets + (
                       b)) < prviousRealTimeInMinuets)) and (nextRealTimeInMinuets is not None) and (nextRealTimeInMinuets is not None):
                           if ((nextExitTimeInMinuets - (b)) > nextRealTimeInMinuets) or ((nextExitTimeInMinuets + (b)) < nextRealTimeInMinuets):

                               # if exit time is bigger - so append to typing arrays none
                               if exitTimeInMinuets > RealTimeInMinuets:

                                   exitTimeARR[c] = None
                                   arrivalTimeARR[c] = None
                                   trainNameARR[c] = None
                                   RealTimeInMinuets = RealTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                                   #clean day of today from RealTime
                                   if d1 in RealTimeInMinuets:
                                       RealTimeInMinuets = RealTimeInMinuets.replace(d1, "")
                                   realTimeARR[d] = RealTimeInMinuets
                                   DateARR[d] = Date[g]
                                   RBARR[d] = RB[g]
                                   stopPointARR[d] = stopPoint[g]
                                   theoTimeARR[d] = theoTime[g]
                                   apllTimeARR[d] = apllTime[g]
                                   if (g < len(realTimeDirection)) and (m < len(exitTimeDirection)):
                                       tdeltaARR[t] = None
                                   status = "yes"
                                   c = c + 1
                                   d = d + 1
                                   g = g + 1





                           # if real time is bigger - so append to franchisee arrays none
                               else:
                                   realTimeARR[d] = None
                                   DateARR[d] = None
                                   RBARR[d] = None
                                   stopPointARR[d] = None
                                   theoTimeARR[d] = None
                                   apllTimeARR[d] = None

                                   if exitTimeInMinuets is not None:
                                       exitTimeInMinuets = exitTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                                       if d1 in exitTimeInMinuets:
                                           exitTimeInMinuets = exitTimeInMinuets.replace(d1, "")
                                   exitTimeARR[c] = exitTimeInMinuets
                                   arrivalTimeARR[c] = arrivalTimeNonBlue[m]
                                   trainNameARR[c] = trainNameNonBlue[m]
                                   if (g < len(realTimeDirection)) and (m < len(exitTimeDirection)):
                                       tdeltaARR[t] = None
                                   c = c + 1
                                   d = d + 1
                                   m = m + 1
                                   status = "t minus one"

                           else:


                               if exitTimeInMinuets is not None:
                                   exitTimeInMinuets = exitTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                                   if d1 in exitTimeInMinuets:
                                       exitTimeInMinuets = exitTimeInMinuets.replace(d1, "")
                               whatToMarkInYellowBackground.append(exitTimeInMinuets)
                               exitTimeARR[c] = (exitTimeInMinuets)
                               arrivalTimeARR[c] = arrivalTimeNonBlue[m]
                               trainNameARR[c] = trainNameNonBlue[m]
                               RealTimeInMinuets = RealTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                               if d1 in RealTimeInMinuets:
                                   RealTimeInMinuets = RealTimeInMinuets.replace(d1, "")
                               realTimeARR[d] = RealTimeInMinuets
                               DateARR[d] = Date[g]
                               RBARR[d] = RB[g]
                               stopPointARR[d] = stopPoint[g]
                               theoTimeARR[d] = theoTime[g]
                               apllTimeARR[d] = apllTime[g]
                               if (g < len(realTimeDirection)) and (m < len(exitTimeDirection)):
                                   tdeltaARR[t] = tdelta
                               m = m + 1
                               status = "no"
                               c = c + 1
                               d = d + 1
                               g = g + 1




                       else:

                               if exitTimeInMinuets is not None:
                                   exitTimeInMinuets = exitTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                                   if d1 in exitTimeInMinuets:
                                       exitTimeInMinuets = exitTimeInMinuets.replace(d1, "")
                               whatToMarkInYellowBackground.append(exitTimeInMinuets)
                               exitTimeARR[c] = (exitTimeInMinuets)
                               arrivalTimeARR[c] = arrivalTimeNonBlue[m]
                               trainNameARR[c] = trainNameNonBlue[m]
                               RealTimeInMinuets = RealTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                               if d1 in RealTimeInMinuets:
                                   RealTimeInMinuets = RealTimeInMinuets.replace(d1, "")
                               realTimeARR[d] = RealTimeInMinuets
                               DateARR[d] = Date[g]
                               RBARR[d] = RB[g]
                               stopPointARR[d] = stopPoint[g]
                               theoTimeARR[d] = theoTime[g]
                               apllTimeARR[d] = apllTime[g]
                               if (g < len(realTimeDirection)) and (m < len(exitTimeDirection)):
                                   tdeltaARR[t] = tdelta
                               m = m + 1
                               status = "no"
                               c = c + 1
                               d = d + 1
                               g = g + 1


                   # if time diff is is proper add values to arrays:
                   else:
                       if exitTimeInMinuets is not None:
                           exitTimeInMinuets = exitTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                           if d1 in exitTimeInMinuets:
                               exitTimeInMinuets = exitTimeInMinuets.replace(d1, "")
                       exitTimeARR[c] = (exitTimeInMinuets)
                       arrivalTimeARR[c] = arrivalTimeNonBlue[m]
                       trainNameARR[c] = trainNameNonBlue[m]
                       RealTimeInMinuets = RealTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                       if d1 in RealTimeInMinuets:
                           RealTimeInMinuets = RealTimeInMinuets.replace(d1, "")
                       realTimeARR[d] = RealTimeInMinuets
                       DateARR[d] = Date[g]
                       RBARR[d] = RB[g]
                       stopPointARR[d] = stopPoint[g]
                       theoTimeARR[d] = theoTime[g]
                       apllTimeARR[d] = apllTime[g]

                       if (g < len(realTimeDirection)) and (m < len(exitTimeDirection)):
                           tdeltaARR[t] = tdelta
                       m = m + 1
                       status = "no"
                       c = c + 1
                       d = d + 1
                       g = g + 1



               #if exit time is none anyway add it to arrays

               else:
                   if exitTimeInMinuets is not None:
                       exitTimeInMinuets = exitTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                       if d1 in exitTimeInMinuets:
                           exitTimeInMinuets = exitTimeInMinuets.replace(d1, "")
                   exitTimeARR[c] = (exitTimeInMinuets)
                   arrivalTimeARR[c] = arrivalTimeNonBlue[m]
                   trainNameARR[c] = trainNameNonBlue[m]
                   if type(RealTimeInMinuets) is not str:
                       RealTimeInMinuets = RealTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                   if d1 in RealTimeInMinuets:
                       RealTimeInMinuets = RealTimeInMinuets.replace(d1, "")
                   realTimeARR[d] = RealTimeInMinuets
                   DateARR[d] = Date[g]
                   RBARR[d] = RB[g]
                   stopPointARR[d] = stopPoint[g]
                   theoTimeARR[d] = theoTime[g]
                   apllTimeARR[d] = apllTime[g]
                   if (g < len(realTimeDirection)) and (m < len(exitTimeDirection)):
                      tdeltaARR[t] = tdelta
                   m = m + 1
                   status = "no"
                   c = c + 1
                   d = d + 1
                   g = g + 1

         else:
               if exitTimeInMinuets is not None and type(exitTimeInMinuets) is not str:
                   exitTimeInMinuets = exitTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                   if d1 in exitTimeInMinuets:
                       exitTimeInMinuets = exitTimeInMinuets.replace(d1, "")
                   elif type(exitTimeInMinuets) is str:
                       exitTimeInMinuets = exitTimeInMinuets.replace(d1, "")

               exitTimeARR[c] = (exitTimeInMinuets)
               arrivalTimeARR[c] = arrivalTimeNonBlue[m]
               trainNameARR[c] = trainNameNonBlue[m]
               if (g < len(realTimeDirection)) and (m < len(exitTimeDirection)):
                   tdeltaARR[t] = tdelta
               if RealTimeInMinuets is not None and type(RealTimeInMinuets) is not str and RealTimeInMinuets != " ":
                   RealTimeInMinuets = RealTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
                   if d1 in RealTimeInMinuets:
                       RealTimeInMinuets = RealTimeInMinuets.replace(d1, "")
               elif RealTimeInMinuets is not None and type(RealTimeInMinuets) is str and RealTimeInMinuets != " ":
                       RealTimeInMinuets = RealTimeInMinuets.replace(d1, "")

               realTimeARR[d] = RealTimeInMinuets
               DateARR[d] = Date[g]
               RBARR[d] = RB[g]
               stopPointARR[d] = stopPoint[g]
               theoTimeARR[d] = theoTime[g]
               apllTimeARR[d] = apllTime[g]
               m = m + 1
               status = "no"
               c = c + 1
               d = d + 1
               g = g + 1


       else:
           exitTimeInMinuets = exitTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
           if d1 in exitTimeInMinuets:
               exitTimeInMinuets = exitTimeInMinuets.replace(d1, "")
           exitTimeARR[c] = (exitTimeInMinuets)
           arrivalTimeARR[c] = arrivalTimeNonBlue[m]
           trainNameARR[c] = trainNameNonBlue[m]
           if (g < len(realTimeDirection)) and (m < len(exitTimeDirection)):
               tdeltaARR[t] = tdelta
           RealTimeInMinuets = RealTimeInMinuets.strftime("%d/%m/%Y %H:%M:%S")
           if d1 in RealTimeInMinuets:
               RealTimeInMinuets = RealTimeInMinuets.replace(d1, "")
           realTimeARR[d] = RealTimeInMinuets
           DateARR[d] = Date[g]
           RBARR[d] = RB[g]
           stopPointARR[d] = stopPoint[g]
           theoTimeARR[d] = theoTime[g]
           apllTimeARR[d] = apllTime[g]

           m = m + 1
           status = "no"
           c = c + 1
           d = d + 1
           g = g + 1



   if south == 0:
       cleanArrFromNeddlessArr(DateARR, redDate, whatToMarkInYellowBackground, sout =0, whereToPaste = 1, s=0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(RBARR, redRB, whatToMarkInYellowBackground, sout = 0, whereToPaste = 2,s=0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(stopPointARR, redStopPoint, whatToMarkInYellowBackground, sout = 0, whereToPaste = 4,s=2, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(theoTimeARR, redTheoTime, whatToMarkInYellowBackground, sout = 0, whereToPaste = 5,s=0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(apllTimeARR, redApllTime, whatToMarkInYellowBackground, sout = 0, whereToPaste = 6, s=0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(realTimeARR, RedRealTimeARR, whatToMarkInYellowBackground, sout =0, whereToPaste =7,s =0, paintInYellowEhereItsNotNone = 1)

       cleanArrFromNeddlessArr(trainNameARR, trainNameBlue, whatToMarkInYellowBackground, sout = 0, whereToPaste=8, s =0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(exitTimeARR, BlueExitTimeARR,  whatToMarkInYellowBackground, sout = 0, whereToPaste=9, s =1, paintInYellowEhereItsNotNone = 0)
       cleanArrFromNeddlessArr(arrivalTimeARR, arrivalTimeBlue, whatToMarkInYellowBackground, sout = 0, whereToPaste=10, s =0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(tdeltaARR, arrivalTimeBlue, whatToMarkInYellowBackground, sout = 0, whereToPaste=11, s=17, paintInYellowEhereItsNotNone = 1)

   else:
       cleanArrFromNeddlessArr(DateARR, redDate, whatToMarkInYellowBackground, sout = 1, whereToPaste=14, s =0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(RBARR, redRB, whatToMarkInYellowBackground, sout = 1, whereToPaste=15, s =0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(stopPointARR, redStopPoint, whatToMarkInYellowBackground, sout =1, whereToPaste=17, s =2, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(theoTimeARR, redTheoTime, whatToMarkInYellowBackground, sout = 1, whereToPaste=18, s =0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(apllTimeARR, redApllTime, whatToMarkInYellowBackground, sout =1, whereToPaste=19, s =0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(realTimeARR, RedRealTimeARR, whatToMarkInYellowBackground, sout =1, whereToPaste=20, s =0, paintInYellowEhereItsNotNone = 1)

       cleanArrFromNeddlessArr(trainNameARR, trainNameBlue, whatToMarkInYellowBackground, sout =1, whereToPaste = 21, s =0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(exitTimeARR, BlueExitTimeARR, whatToMarkInYellowBackground, sout =1, whereToPaste = 22, s =1, paintInYellowEhereItsNotNone = 0)
       cleanArrFromNeddlessArr(arrivalTimeARR, arrivalTimeBlue, whatToMarkInYellowBackground, sout =1 , whereToPaste =23, s =0, paintInYellowEhereItsNotNone = 1)
       cleanArrFromNeddlessArr(tdeltaARR, arrivalTimeBlue, whatToMarkInYellowBackground, sout =1, whereToPaste=24, s=17, paintInYellowEhereItsNotNone = 1)


   return


def cleanArrFromNeddlessArr(nonColoredArr, coloredArr, whatToMarkInYellowBackground, sout, whereToPaste, s, paintInYellowEhereItsNotNone):
   arrForNonColoredArr = []
   numberOfNones = 0
   for i in range(0, len(nonColoredArr) - 1):
       nonColoredArr[i]
       if nonColoredArr[i] is not None:
           theBiggestINdexThatIsNotEmpty = i
   for w in range(0, len(coloredArr)):
       if coloredArr[w] is not None and coloredArr[w] != 'None':
           arrForNonColoredArr.append(coloredArr[w])

       else:
           numberOfNones = numberOfNones + 1
   for i in range(0, len(arrForNonColoredArr) - 1):
       if arrForNonColoredArr[i] is not None:
           theBiggestINdexThatIsNotEmptyColoredArr = i

   if numberOfNones == 350 or numberOfNones ==349:
       theBiggestINdexThatIsNotEmptyColoredArr = 0
   nonColoredArr = nonColoredArr[0: theBiggestINdexThatIsNotEmpty+1]
   coloredArr = arrForNonColoredArr[0 : theBiggestINdexThatIsNotEmptyColoredArr+2]
   if sout == 0:
       direction =0
   else:
       direction =1

   combineArraysWithDifferentColorsAndThenPasteIt(nonColoredArr, coloredArr, whereToPaste, s,
                                                  theBiggestINdexThatIsNotEmpty,
                                                  whatToMarkInYellowBackground, paintInYellowEhereItsNotNone, sout = direction)


   return(nonColoredArr, coloredArr, sout)

def combineArraysWithDifferentColorsAndThenPasteIt(nonColoredArr, coloredArr, whereToPaste, s,
                                                      theBiggestINdexThatIsNotEmpty, whatToMarkInYellowBackground,
                                                      paintInYellowEhereItsNotNone, sout):
   theBiggestINdexThatIsNotEmpty = theBiggestINdexThatIsNotEmpty

   PasteAfterDiffTimeCalculation(nonColoredArr,  coloredArr, whereToPaste, whatToMarkInYellowBackground,  paintInYellowEhereItsNotNone, s, theBiggestINdexThatIsNotEmpty)
   fileAfterMonitoring2.save(fileNameGenerating + '.xlsx')
   if sout == 0:
       directio =0
   else:
       directio =1

   if s == 1:
       paintInBlueOrRed(coloredArr, nonColoredArr, s,  oneOrTWo = directio, a=1)
   elif s == 2:
       paintInBlueOrRed(coloredArr, nonColoredArr, s, oneOrTWo = directio, a =2)


   return

def PasteAfterDiffTimeCalculation(nonColoredArr, coloredArr, whereToPaste, whatToMarkInYellowBackground, paintInYellowEhereItsNotNone, s, theBiggestINdexThatIsNotEmpty):

  if whereToPaste != 11 and whereToPaste !=24:

      if s != 17:
       nonColoredArr = nonColoredArr + coloredArr
      if CounterHowMuchRedIsThere > 0:
        nonColoredArr = nonColoredArr + ArrCounterHowMuchRedIsThere



      for o in range(0, len(nonColoredArr)):

          if (nonColoredArr[o] is None) and (o < theBiggestINdexThatIsNotEmpty+1):

              if (whereToPaste) < 11:
                  for u in range(1,11):

                    fileAfterMonitoring2[x].cell(row=o + 4, column=u).fill = PatternFill(start_color='FFFFFF00',end_color='FFFFFF00',fill_type='solid')
                    fileAfterMonitoring2[x].cell(row=o + 5, column=u).fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00',fill_type='solid')
                    fileAfterMonitoring2[x].cell(row=o + 6, column=u).fill = PatternFill(start_color='FFFFFF00',end_color='FFFFFF00',fill_type='solid')
              else:
                for u in range(14, 24):
                    fileAfterMonitoring2[x].cell(row=o + 4, column=u).fill = PatternFill(start_color='FFFFFF00',end_color='FFFFFF00', fill_type='solid')
                    fileAfterMonitoring2[x].cell(row=o + 5, column=u).fill = PatternFill(start_color='FFFFFF00',end_color='FFFFFF00',fill_type='solid')
                    fileAfterMonitoring2[x].cell(row=o + 6, column=u).fill = PatternFill(start_color='FFFFFF00',end_color='FFFFFF00',fill_type='solid')


              # fileAfterMonitoring2[x].cell(row=o + 5, column=whereToPaste).fill = PatternFill(start_color='FFFFFF00',
              #                                                                                 end_color='FFFFFF00',
              #                                                                                 fill_type='solid')
              #
              # fileAfterMonitoring2[x].cell(row=o + 6, column=whereToPaste).fill = PatternFill(start_color='FFFFFF00',
              #                                                                                 end_color='FFFFFF00',
              #                                                                                 fill_type='solid')
          if s == 1:
              if (nonColoredArr[o] is not None) and nonColoredArr[o]!= ("None") and nonColoredArr[o] != " ":
                 if ((type(nonColoredArr[o])) is datetime.time):
                      nonColoredArr[o] = nonColoredArr[o].strftime("%d/%m/%Y %H:%M:%S")
                 if d1 in nonColoredArr[o]:
                      nonColoredArr[o] = nonColoredArr[o].replace(d1, "")
          if nonColoredArr[o] != ("None"):
              fileAfterMonitoring2[x].cell(row=o + 5, column=whereToPaste).value = nonColoredArr[o]

          moreThan23hours = datetime.timedelta(hours=23)
          if paintInYellowEhereItsNotNone==0:
              if len(whatToMarkInYellowBackground)>0:
                  for v in range(0,len(whatToMarkInYellowBackground)):
                      if whatToMarkInYellowBackground[v] is not None and fileAfterMonitoring2[x].cell(row=o + 5, column=whereToPaste).value is not None:
                            if whatToMarkInYellowBackground[v] == fileAfterMonitoring2[x].cell(row=o + 5, column=whereToPaste).value:

                                for i in range(-8,2):
                                    exitTimeForCalcualtionOfMoreThan23Hours = fileAfterMonitoring2[x].cell(row=o + 5, column=whereToPaste).value
                                    realTimeForCalcualtionOfMoreThan23Hours = fileAfterMonitoring2[x].cell(row=o + 5, column=whereToPaste-2).value

                                    checkIfNeedToHavingDateOforTodayOrForTommorowRealTime = 0
                                    checkIfNeedToHavingDateOforTodayOrForTommorowExitTime = 0
                                    if type(exitTimeForCalcualtionOfMoreThan23Hours) is str:
                                        if " "  in exitTimeForCalcualtionOfMoreThan23Hours:
                                            exitTimeForCalcualtionOfMoreThan23Hours= exitTimeForCalcualtionOfMoreThan23Hours.replace(" ", "")
                                        if "30/12/1899" in exitTimeForCalcualtionOfMoreThan23Hours:
                                            exitTimeForCalcualtionOfMoreThan23Hours = exitTimeForCalcualtionOfMoreThan23Hours.replace("30/12/1899","")
                                        if "01/01/1900" in exitTimeForCalcualtionOfMoreThan23Hours:
                                            exitTimeForCalcualtionOfMoreThan23Hours = exitTimeForCalcualtionOfMoreThan23Hours("01/01/1900", "")
                                        if "00:" in exitTimeForCalcualtionOfMoreThan23Hours or "01:" in exitTimeForCalcualtionOfMoreThan23Hours or "02:" in exitTimeForCalcualtionOfMoreThan23Hours:
                                            for p in range(0, 6):
                                                p = str(p)
                                                for j in range(0, 10):
                                                    j = str(j)
                                                    for k in range(0, 6):
                                                        k = str(k)
                                                        for m in range(0, 10):
                                                            m = str(m)

                                                            if (
                                                                    "00:" + p + j + ":" + k + m) in exitTimeForCalcualtionOfMoreThan23Hours or (
                                                                    "01:" + p + j + ":" + k + m) in exitTimeForCalcualtionOfMoreThan23Hours or (
                                                                    "02:" + p + j + ":" + k + m) in exitTimeForCalcualtionOfMoreThan23Hours:
                                                                checkIfNeedToHavingDateOforTodayOrForTommorowExitTime = 1


                                    if type(realTimeForCalcualtionOfMoreThan23Hours) is str:
                                        if " " in realTimeForCalcualtionOfMoreThan23Hours:
                                            realTimeForCalcualtionOfMoreThan23Hours = realTimeForCalcualtionOfMoreThan23Hours.replace(" ", "")
                                        if "30/12/1899" in realTimeForCalcualtionOfMoreThan23Hours:
                                            realTimeForCalcualtionOfMoreThan23Hours = realTimeForCalcualtionOfMoreThan23Hours.replace("30/12/1899")
                                        if "01/01/1900" in realTimeForCalcualtionOfMoreThan23Hours:
                                            realTimeForCalcualtionOfMoreThan23Hours = realTimeForCalcualtionOfMoreThan23Hours("01/01/1900", "")
                                        if "00:" in realTimeForCalcualtionOfMoreThan23Hours or "01:" in realTimeForCalcualtionOfMoreThan23Hours or "02:" in realTimeForCalcualtionOfMoreThan23Hours:
                                            for p in range(0, 6):
                                                p = str(p)
                                                for j in range(0, 10):
                                                    j = str(j)
                                                    for k in range(0, 6):
                                                        k = str(k)
                                                        for m in range(0, 10):
                                                            m = str(m)

                                                            if (
                                                                    "00:" + p + j + ":" + k + m) in realTimeForCalcualtionOfMoreThan23Hours or (
                                                                    "01:" + p + j + ":" + k + m) in realTimeForCalcualtionOfMoreThan23Hours or (
                                                                    "02:" + p + j + ":" + k + m) in realTimeForCalcualtionOfMoreThan23Hours:
                                                                checkIfNeedToHavingDateOforTodayOrForTommorowRealTime = 1


                                    tomorrow = today + datetime.timedelta(days=1)


                                    #exitTimeForCalcualtionOfMoreThan23Hours = datetime.datetime.strptime(exitTimeForCalcualtionOfMoreThan23Hours, '%H:%M:%S')
                                    #realTimeForCalcualtionOfMoreThan23Hours = datetime.datetime.strptime(realTimeForCalcualtionOfMoreThan23Hours, '%H:%M:%S')

                                    if type(d1) is not str:
                                        d2 = d1.strftime("%d/%m/%Y")
                                    else:
                                        d2=d1
                                    if type(tomorrow) is not str:
                                        tomorrow2 = tomorrow.strftime("%d/%m/%Y")
                                    else:
                                        tomorrow2 = tomorrow

                                    if checkIfNeedToHavingDateOforTodayOrForTommorowExitTime ==0:
                                        exitTimeForCalcualtionOfMoreThan23Hours = datetime.datetime.strptime(
                                            d2 + " " + exitTimeForCalcualtionOfMoreThan23Hours, "%d/%m/%Y %H:%M:%S")
                                    elif checkIfNeedToHavingDateOforTodayOrForTommorowExitTime ==1:
                                        exitTimeForCalcualtionOfMoreThan23Hours = datetime.datetime.strptime(
                                            tomorrow2 + " " + exitTimeForCalcualtionOfMoreThan23Hours, "%d/%m/%Y %H:%M:%S")

                                    if checkIfNeedToHavingDateOforTodayOrForTommorowRealTime ==0:
                                        realTimeForCalcualtionOfMoreThan23Hours = datetime.datetime.strptime(
                                            d2 + " " + realTimeForCalcualtionOfMoreThan23Hours, "%d/%m/%Y %H:%M:%S")
                                    elif checkIfNeedToHavingDateOforTodayOrForTommorowRealTime ==1:
                                        realTimeForCalcualtionOfMoreThan23Hours = datetime.datetime.strptime(
                                            tomorrow2 + " " + realTimeForCalcualtionOfMoreThan23Hours, "%d/%m/%Y %H:%M:%S")

                                    if exitTimeForCalcualtionOfMoreThan23Hours - moreThan23hours < realTimeForCalcualtionOfMoreThan23Hours and realTimeForCalcualtionOfMoreThan23Hours - moreThan23hours < exitTimeForCalcualtionOfMoreThan23Hours:
                                        fileAfterMonitoring2[x].cell(row=o + 5, column=whereToPaste+i).fill=PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type = 'solid')

  else:
      settingTimedeltaAsAFormula(nonColoredArr, whereToPaste)





  return

def settingTimedeltaAsAFormula(nonColoredArr, whereToPaste):
    print("I'm in settingTimedeltaAsAFormula")

    itsAfterMidNight = 0
    forExitTimeAfterMIdNight = 0
    forRealTimeAfterMIdNight = 0

    tomorrow = today + datetime.timedelta(days=1)
    if type(d1) is not str:
        d2 = d1.strftime("%d/%m/%Y")
    else:
        d2 = d1
    if type(tomorrow) is not str:
        tomorrow2 = tomorrow.strftime("%d/%m/%Y")
    else:
        tomorrow2 = tomorrow

    for m in range(5, 250):

        if (fileAfterMonitoring2[x].cell(row=m, column=whereToPaste-2).value is not None) or (fileAfterMonitoring2[x].cell(row=m, column=whereToPaste-4).value) is not None:
            m = str(m)
            if whereToPaste==11:
                cellForPastingFormula = "K" + m
                cellExitTime = "I" + m
                cellRealTime = "G" + m
            else:
                cellForPastingFormula = "X" + m
                cellExitTime = "V" + m
                cellRealTime = "T" + m
            formula = "=IF(" + cellExitTime + "-" + cellRealTime + ">0," + cellExitTime + "-" + cellRealTime + "," + cellRealTime + "-" + cellExitTime + ")"
            if type(fileAfterMonitoring2[x][cellRealTime].value) is str and (fileAfterMonitoring2[x][cellRealTime].value) is not None:
                if (fileAfterMonitoring2[x][cellRealTime].value) is not None:
                    doesRealTimeIsAfterMidNight = fileAfterMonitoring2[x][cellRealTime].value
                    if "30/12/1899" in doesRealTimeIsAfterMidNight:
                        doesRealTimeIsAfterMidNight = doesRealTimeIsAfterMidNight.replace("30/12/1899", "")
                    if " " in doesRealTimeIsAfterMidNight:
                        doesRealTimeIsAfterMidNight = doesRealTimeIsAfterMidNight.replace(" ", "")

                    if doesRealTimeIsAfterMidNight is not None and len(doesRealTimeIsAfterMidNight)>0:
                        doesRealTimeIsAfterMidNight = datetime.datetime.strptime(doesRealTimeIsAfterMidNight, "%H:%M:%S")


            else:
                if (fileAfterMonitoring2[x][cellRealTime].value) is not None:
                    doesRealTimeIsAfterMidNight = fileAfterMonitoring2[x][cellRealTime].value.strftime("%H:%M:%S")

            if type(doesRealTimeIsAfterMidNight) is not str:
                if doesRealTimeIsAfterMidNight is not None:
                    checkingAfterMidNightOrNot = doesRealTimeIsAfterMidNight.strftime("%H:%M:%S")
            else:
                checkingAfterMidNightOrNot = doesRealTimeIsAfterMidNight

            if "00:" in checkingAfterMidNightOrNot or "01:" in checkingAfterMidNightOrNot or "02:" in checkingAfterMidNightOrNot:
                for p in range(0, 6):
                    p = str(p)
                    for j in range(0, 10):
                        j = str(j)
                        for k in range(0, 6):
                            k = str(k)
                            for m in range(0, 10):
                                m = str(m)

                                if ("00:" + p + j + ":" + k + m) in checkingAfterMidNightOrNot or (
                                        "01:" + p + j + ":" + k + m) in checkingAfterMidNightOrNot or (
                                        "02:" + p + j + ":" + k + m) in checkingAfterMidNightOrNot:
                                    itsAfterMidNight = 1
                                    forExitTimeAfterMIdNight = 1



            if type(fileAfterMonitoring2[x][cellExitTime].value) is str:
                doesExitTimeIsAfterMidNight = fileAfterMonitoring2[x][cellExitTime].value
                if "30/12/1899" in doesExitTimeIsAfterMidNight:
                    doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace("30/12/1899","")
                if " " in doesExitTimeIsAfterMidNight and len(doesExitTimeIsAfterMidNight)>1:
                    doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace(" ", "")
                    doesExitTimeIsAfterMidNight = datetime.datetime.strptime(doesExitTimeIsAfterMidNight, "%H:%M:%S")
                else:
                    doesExitTimeIsAfterMidNight = None
            else:
                if fileAfterMonitoring2[x][cellExitTime].value is not None:
                    doesExitTimeIsAfterMidNight = fileAfterMonitoring2[x][cellExitTime].value.strftime("%m/%d/%Y, %H:%M:%S")
                    if "30/12/1899" in doesExitTimeIsAfterMidNight:
                        doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace("30/12/1899", "")

            if type(doesExitTimeIsAfterMidNight) is not str and doesExitTimeIsAfterMidNight is not None:
                checkingAfterMidNightOrNot = doesExitTimeIsAfterMidNight.strftime("%H:%M:%S")
            else:
                checkingAfterMidNightOrNot = doesExitTimeIsAfterMidNight
            if checkingAfterMidNightOrNot is not None:

                if "00:" in checkingAfterMidNightOrNot or "01:" in checkingAfterMidNightOrNot or "02:" in checkingAfterMidNightOrNot:
                    for p in range(0, 6):
                        p = str(p)
                        for j in range(0, 10):
                            j = str(j)
                            for k in range(0, 6):
                                k = str(k)
                                for m in range(0, 10):
                                    m = str(m)
                                    if ("00:" + p + j + ":" + k + m) in checkingAfterMidNightOrNot or (
                                            "01:" + p + j + ":" + k + m) in checkingAfterMidNightOrNot or (
                                            "02:" + p + j + ":" + k + m) in checkingAfterMidNightOrNot:
                                        itsAfterMidNight = 1
                                        forRealTimeAfterMIdNight = 1


            if forExitTimeAfterMIdNight == 1 and forRealTimeAfterMIdNight == 0:
                if (type(doesRealTimeIsAfterMidNight)) is str:
                    if ("01/01/1900" in doesRealTimeIsAfterMidNight):
                        doesRealTimeIsAfterMidNight = doesRealTimeIsAfterMidNight.replace("01/01/1900 ", "")
                    else:
                        doesRealTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellRealTime].value)
                else:
                    doesRealTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellRealTime].value)
                    # if ("01/01/1900" in doesExitTimeIsAfterMidNight):
                    #     doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace("01/01/1900 ", " ")

                if (type(doesExitTimeIsAfterMidNight)) is str:
                    if ("01/01/1900" in doesExitTimeIsAfterMidNight):
                        doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace("01/01/1900 ", " ")
                    else:
                        doesExitTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellExitTime].value)
                        if doesExitTimeIsAfterMidNight is not None:
                            if ("01/01/1900" in doesExitTimeIsAfterMidNight):
                                doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace("01/01/1900 ", " ")
                else:
                    doesExitTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellExitTime].value)


                if type(doesRealTimeIsAfterMidNight) is not str:
                    if doesRealTimeIsAfterMidNight is not None:
                        doesRealTimeIsAfterMidNight = doesRealTimeIsAfterMidNight.strftime("%H:%M:%S")
                if type(doesExitTimeIsAfterMidNight) is not str:
                    if (doesExitTimeIsAfterMidNight) is not None:
                        doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.strftime("%H:%M:%S")


                if (doesRealTimeIsAfterMidNight) is not None and len(doesRealTimeIsAfterMidNight) > 1:
                    fileAfterMonitoring2[x][cellRealTime].value = datetime.datetime.strptime(tomorrow2 + " " + doesRealTimeIsAfterMidNight,"%d/%m/%Y %H:%M:%S")
                if (doesExitTimeIsAfterMidNight) is not None and len(doesExitTimeIsAfterMidNight) > 1:
                    fileAfterMonitoring2[x][cellExitTime].value = datetime.datetime.strptime(d2 + " " + doesExitTimeIsAfterMidNight,"%d/%m/%Y %H:%M:%S")

            elif forExitTimeAfterMIdNight == 1 and forRealTimeAfterMIdNight == 1:

                if (type(fileAfterMonitoring2[x][cellRealTime].value)) is not str:
                    if ("01/01/1900" in doesRealTimeIsAfterMidNight):
                        doesRealTimeIsAfterMidNight = doesRealTimeIsAfterMidNight.replace("01/01/1900 ", "")
                    else:
                        doesRealTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellRealTime].value)

                else:
                    doesRealTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellRealTime].value)
                    if "01/01/1900" in doesRealTimeIsAfterMidNight:
                        doesRealTimeIsAfterMidNight = doesRealTimeIsAfterMidNight.replace("01/01/1900","")
                    if "30/12/1899" in doesRealTimeIsAfterMidNight:
                        doesRealTimeIsAfterMidNight = doesRealTimeIsAfterMidNight.replace("30/12/1899","")

                if (type(fileAfterMonitoring2[x][cellExitTime].value)) is not str:

                    if ("01/01/1900" in doesExitTimeIsAfterMidNight):
                        doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace("01/01/1900 ", "")
                    else:
                        doesExitTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellExitTime].value)
                else:

                    doesExitTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellExitTime].value)
                    if "01/01/1900" in doesExitTimeIsAfterMidNight:
                        doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace("01/01/1900","")
                    if "30/12/1899" in doesExitTimeIsAfterMidNight:
                        doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace("30/12/1899","")

                fileAfterMonitoring2[x][cellRealTime].value = datetime.datetime.strptime(tomorrow2 + " " + doesRealTimeIsAfterMidNight,"%d/%m/%Y %H:%M:%S")
                fileAfterMonitoring2[x][cellExitTime].value = datetime.datetime.strptime(tomorrow2 + " " + doesExitTimeIsAfterMidNight,"%d/%m/%Y %H:%M:%S")


            elif forExitTimeAfterMIdNight == 0 and forRealTimeAfterMIdNight == 1:

                if (type(fileAfterMonitoring2[x][cellRealTime].value)) is not str:
                    if ("01/01/1900" in doesRealTimeIsAfterMidNight):
                        doesRealTimeIsAfterMidNight = doesRealTimeIsAfterMidNight.replace("01/01/1900 ", "")
                    else:
                        doesRealTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellRealTime].value)
                else:
                    doesRealTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellRealTime].value)

                if (type(fileAfterMonitoring2[x][cellExitTime].value)) is not str:
                    if ("01/01/1900" in doesExitTimeIsAfterMidNight):
                        doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace("01/01/1900 ", "")
                    else:
                        doesExitTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellExitTime].value)
                else:
                    doesExitTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellExitTime].value)
                if (doesRealTimeIsAfterMidNight) is not None and len(doesRealTimeIsAfterMidNight) > 1:
                    fileAfterMonitoring2[x][cellRealTime].value = datetime.datetime.strptime(d2 + " " + doesRealTimeIsAfterMidNight,"%d/%m/%Y %H:%M:%S")
                if (doesExitTimeIsAfterMidNight) is not None and len(doesExitTimeIsAfterMidNight) > 1:
                    fileAfterMonitoring2[x][cellExitTime].value = datetime.datetime.strptime(tomorrow2 + " " + doesExitTimeIsAfterMidNight,"%d/%m/%Y %H:%M:%S")

            else:
                if (type(fileAfterMonitoring2[x][cellRealTime].value)) is not str and (fileAfterMonitoring2[x][cellRealTime].value) is not None:
                    if ("01/01/1900" in doesRealTimeIsAfterMidNight):
                        doesRealTimeIsAfterMidNight = doesRealTimeIsAfterMidNight.replace("01/01/1900 ", "")
                    else:
                        doesRealTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellRealTime].value)
                else:
                    doesRealTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellRealTime].value)

                if (type(fileAfterMonitoring2[x][cellExitTime].value)) is not str and (fileAfterMonitoring2[x][cellExitTime].value) is not None:
                    if ("01/01/1900" in doesExitTimeIsAfterMidNight):
                        doesExitTimeIsAfterMidNight = doesExitTimeIsAfterMidNight.replace("01/01/1900 ", "")
                    else:
                        doesExitTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellExitTime].value)
                else:
                    doesExitTimeIsAfterMidNight = (fileAfterMonitoring2[x][cellExitTime].value)

                if (doesRealTimeIsAfterMidNight) is not None and len(doesRealTimeIsAfterMidNight) > 1:
                    fileAfterMonitoring2[x][cellRealTime].value = datetime.datetime.strptime(d2 + " " + doesRealTimeIsAfterMidNight,"%d/%m/%Y %H:%M:%S")
                if (doesExitTimeIsAfterMidNight) is not None and len(doesExitTimeIsAfterMidNight) > 1:
                    fileAfterMonitoring2[x][cellExitTime].value = datetime.datetime.strptime(d2 + " " + doesExitTimeIsAfterMidNight, "%d/%m/%Y %H:%M:%S")


            fileAfterMonitoring2[x][cellForPastingFormula] = formula

    if x== earliest:
        date_style = NamedStyle(name='date_style', number_format='HH:MM:SS')
        wb = openpyxl.Workbook()
        wb.add_named_style(date_style)
        wb.save(r'C:\Users\shoshana\PycharmProjects\pandas\Style_1.xlsx')
        wb['Sheet']['A1'].style = date_style
        wb['Sheet'].style = 'date_style'
        wb.save(r'C:\Users\shoshana\PycharmProjects\pandas\Style_1.xlsx')
    else:
        wb= openpyxl.load_workbook(r'C:\Users\shoshana\PycharmProjects\pandas\Style_1.xlsx')



    for m in range(5, 250):

        if (fileAfterMonitoring2[x].cell(row=m, column=whereToPaste-2).value is not None) or (fileAfterMonitoring2[x].cell(row=m, column=whereToPaste-4).value is not None):
            m = str(m)
            if whereToPaste == 11:
                cellExitTime = "I" + m
                cellRealTime = "G" + m
                formulaCell = "K" + m
            else:
                cellExitTime = "V" + m
                cellRealTime = "T" + m
                formulaCell = "X" + m


            fileAfterMonitoring2[x][cellExitTime].number_format = copy(wb['Sheet']['A1'].number_format)
            fileAfterMonitoring2[x][cellRealTime].number_format = copy(wb['Sheet']['A1'].number_format)
            fileAfterMonitoring2[x][formulaCell].number_format = copy(wb['Sheet']['A1'].number_format)


    return

def paintInBlueOrRed(coloredArr, nonColoredArr,s,  oneOrTWo, a):
  nonColoredArr = nonColoredArr + coloredArr
  print(" I'm into paintInBlueOrRed")
  decideIfToPaintOrNot = 0
  counterNumberOfPaints = 0

  for j in range(0, len(coloredArr)):
      if coloredArr[j] is None or coloredArr[j] == "None":
          decideIfToPaintOrNot = decideIfToPaintOrNot + 0
      else:
          decideIfToPaintOrNot = decideIfToPaintOrNot + 3
          counterNumberOfPaints = counterNumberOfPaints + 1


  for l in range (0, counterNumberOfPaints):

      if oneOrTWo ==0:
       if a == 1 and decideIfToPaintOrNot > 0:

           fileAfterMonitoring2[x].cell(row= len(nonColoredArr) + 4 - l, column=9).font = Font(color="FF00B0F0")





       elif a == 2 and decideIfToPaintOrNot > 0:
          fileAfterMonitoring2[x].cell(row=len(nonColoredArr) +4 -l, column=4).font = Font(color="FFFF0000")


      else:
       if a == 1 and decideIfToPaintOrNot > 0:
           fileAfterMonitoring2[x].cell(row= len(nonColoredArr) +4 - l, column=22).font = Font(color="FF00B0F0")

       elif a == 2 and decideIfToPaintOrNot > 0:
          fileAfterMonitoring2[x].cell(row=len(nonColoredArr) +4 -l, column=17).font = Font(color="FFFF0000")

  return


def paintInOrangeAllTheCellThatWasPainted(exitTimeColumn, arrivalTimeColumn):
    global doINeedToPaintToOrangeInEXitTimeColumnSouth
    doINeedToPaintToOrangeInEXitTimeColumnSouth = 0
    global doINeedToPaintToOrangeInArrivalTimeColumnSouth
    doINeedToPaintToOrangeInArrivalTimeColumnSouth = 0
    global orangedExitTimeArrSouth
    orangedExitTimeArrSouth = []
    global orangedArrivalTimeArrSouth
    orangedArrivalTimeArrSouth = []

    global doINeedToPaintToOrangeInEXitTimeColumnNorth
    doINeedToPaintToOrangeInEXitTimeColumnNorth = 0
    global doINeedToPaintToOrangeInArrivalTimeColumnNorth
    doINeedToPaintToOrangeInArrivalTimeColumnNorth = 0

    global orangedExitTimeArrNorth
    orangedExitTimeArrNorth = []
    global orangedArrivalTimeArrNorth
    orangedArrivalTimeArrNorth = []

    for l in range(0, 250):  # go over all cells
        if exitTimeColumn ==9:
            if (book[x].cell(row=l + 5, column=exitTimeColumn).value) is not None:
                try:
                    if (book[x].cell(row=l + 5, column=exitTimeColumn).font.color.rgb) == 'FFFFC000':
                        orangedExitTimeArrSouth.append(book[x].cell(row=l + 5, column=exitTimeColumn).value)
                        doINeedToPaintToOrangeInEXitTimeColumnSouth = doINeedToPaintToOrangeInEXitTimeColumnSouth + 1
                except Exception:
                    pass

            if (book[x].cell(row=l + 5, column=arrivalTimeColumn).value) is not None:
                try:
                    if (book[x].cell(row=l + 5, column=arrivalTimeColumn).font.color.rgb) == 'FFFFC000':
                        orangedArrivalTimeArrSouth.append(book[x].cell(row=l + 5, column=arrivalTimeColumn).value)
                        doINeedToPaintToOrangeInArrivalTimeColumnSouth = doINeedToPaintToOrangeInArrivalTimeColumnSouth + 1

                except Exception:
                    pass

        elif exitTimeColumn == 22:
            if (book[x].cell(row=l + 5, column=exitTimeColumn).value) is not None:
                try:
                    if (book[x].cell(row=l + 5, column=exitTimeColumn).font.color.rgb) == 'FFFFC000':
                        orangedExitTimeArrNorth.append(book[x].cell(row=l + 5, column=exitTimeColumn).value)
                        doINeedToPaintToOrangeInEXitTimeColumnNorth = doINeedToPaintToOrangeInEXitTimeColumnNorth + 1
                except Exception:
                    pass

        if (book[x].cell(row=l + 5, column=arrivalTimeColumn).value) is not None:
            try:
                if (book[x].cell(row=l + 5, column=arrivalTimeColumn).font.color.rgb) == 'FFFFC000':
                    orangedArrivalTimeArrNorth.append(book[x].cell(row=l + 5, column=arrivalTimeColumn).value)
                    doINeedToPaintToOrangeInArrivalTimeColumnNorth = doINeedToPaintToOrangeInArrivalTimeColumnNorth + 1
            except Exception:
                pass

    if doINeedToPaintToOrangeInEXitTimeColumnSouth > 0:
        for d in range(0,len(orangedExitTimeArrSouth)):
            for u in range (0,250):
                if fileAfterMonitoring2[x].cell(row=u +5 , column=9).value is not None:
                    if orangedExitTimeArrSouth[d] == fileAfterMonitoring2[x].cell(row=u +5 , column=9).value:
                        fileAfterMonitoring2[x].cell(row=u + 5, column=9).font = Font(color="FFFFC000")



    if doINeedToPaintToOrangeInArrivalTimeColumnSouth > 0:

        for d in range(0,len(orangedArrivalTimeArrSouth)):

            for u in range (0,250):
                if fileAfterMonitoring2[x].cell(row=u +5 , column=10).value is not None:
                    if orangedArrivalTimeArrSouth[d] == fileAfterMonitoring2[x].cell(row=u +5 , column=10).value:
                        fileAfterMonitoring2[x].cell(row=u + 5, column=10).font = Font(color="FFFFC000")

    if doINeedToPaintToOrangeInEXitTimeColumnNorth > 0:
        for d in range(0,len(orangedExitTimeArrNorth)):
            for u in range (0,250):
                if fileAfterMonitoring2[x].cell(row=u +5 , column=22).value is not None:
                    if orangedExitTimeArrNorth[d] == fileAfterMonitoring2[x].cell(row=u +5 , column=22).value:
                        fileAfterMonitoring2[x].cell(row=u + 5, column=22).font = Font(color="FFFFC000")



    if doINeedToPaintToOrangeInArrivalTimeColumnNorth > 0:
        for d in range(0,len(orangedArrivalTimeArrNorth)):
            for u in range (0,250):
                if fileAfterMonitoring2[x].cell(row=u +5 , column=23).value is not None:
                    if orangedArrivalTimeArrNorth[d] == fileAfterMonitoring2[x].cell(row=u +5 , column=10).value:
                        fileAfterMonitoring2[x].cell(row=u + 5, column=23).font = Font(color="FFFFC000")

    return

def AddingHeadLinesfromFranchiser():
   fromWhereToTake = [1, 2, 4, 5, 6 , 7, 8, 9, 10, 14, 15 , 17, 18, 19, 20, 21, 22, 23]
   whereToPaste = [1, 2,  4, 5, 6 , 7, 8, 9, 10, 14, 15 , 17, 18, 19, 20, 21, 22, 23]
   #style = [font, border, fill, number_format, protection, alignment]
   for i in range(0, 18):
           fileAfterMonitoring2[x].cell(row=4, column=1).value = book[x].cell(row=4, column=1).value
           fileAfterMonitoring2[x].cell(row=4, column=2).value = book[x].cell(row=4, column=2).value
           fileAfterMonitoring2[x].cell(row=4, column=4).value = book[x].cell(row=4, column=4).value
           fileAfterMonitoring2[x].cell(row=4, column=5).value = book[x].cell(row=4, column=5).value
           fileAfterMonitoring2[x].cell(row=4, column=6).value = book[x].cell(row=4, column=6).value
           fileAfterMonitoring2[x].cell(row=4, column=7).value = book[x].cell(row=4, column=7).value
           fileAfterMonitoring2[x].cell(row=4, column=8).value = book[x].cell(row=4, column=8).value
           fileAfterMonitoring2[x].cell(row=4, column=9).value = book[x].cell(row=4, column=9).value
           fileAfterMonitoring2[x].cell(row=4, column=10).value = book[x].cell(row=4, column=10).value

           fileAfterMonitoring2[x].cell(row=4, column=14).value = book[x].cell(row=4, column=14).value
           fileAfterMonitoring2[x].cell(row=4, column=15).value = book[x].cell(row=4, column=15).value
           fileAfterMonitoring2[x].cell(row=4, column=17).value = book[x].cell(row=4, column=17).value
           fileAfterMonitoring2[x].cell(row=4, column=18).value = book[x].cell(row=4, column=18).value
           fileAfterMonitoring2[x].cell(row=4, column=19).value = book[x].cell(row=4, column=19).value
           fileAfterMonitoring2[x].cell(row=4, column=20).value = book[x].cell(row=4, column=20).value
           fileAfterMonitoring2[x].cell(row=4, column=21).value = book[x].cell(row=4, column=21).value
           fileAfterMonitoring2[x].cell(row=4, column=22).value = book[x].cell(row=4, column=22).value
           fileAfterMonitoring2[x].cell(row=4, column=23).value = book[x].cell(row=4, column=23).value


           fileAfterMonitoring2[x].cell(row=4, column= whereToPaste[i]).font = copy(book[x].cell(row=4, column= fromWhereToTake[i]).font)
           fileAfterMonitoring2[x].cell(row=4, column= whereToPaste[i]).border = copy(book[x].cell(row=4, column= fromWhereToTake[i]).border)
           fileAfterMonitoring2[x].cell(row=4, column= whereToPaste[i]).fill = copy(book[x].cell(row=4, column= fromWhereToTake[i]).fill)
           fileAfterMonitoring2[x].cell(row=4, column= whereToPaste[i]).number_format = copy(book[x].cell(row=4, column= fromWhereToTake[i]).number_format)
           fileAfterMonitoring2[x].cell(row=4, column= whereToPaste[i]).protection = copy(book[x].cell(row=4, column= fromWhereToTake[i]).protection)
           fileAfterMonitoring2[x].cell(row=4, column= whereToPaste[i]).alignment = copy(book[x].cell(row=4, column= fromWhereToTake[i]).alignment)

   return

def finalPating():
    TypingReportLoad = openpyxl.load_workbook(r"C:\Users\shoshana\PycharmProjects\pandas\TypingReportSeptember.xlsx")
    TypingReport = TypingReportLoad["Sheet1"]
    b = 2
    for o in range(2, 10000):

        if (not ((TypingReport.cell(row=o, column=1).value) is None)):
            dateCellValue = TypingReport.cell(row=o, column=1).value
            dateCellValue = dateCellValue.strftime('%m/%d/%Y')
            if (o != 2):
                if o > 4:
                    sheetPrevious = sheet
                if ((dateCellValue) != ((TypingReport.cell(row=o - 1, column=1).value).strftime('%m/%d/%Y'))):
                    if ((dateCellValue[3] != '0') and not ((dateCellValue[4] != '1'))):
                        b = 2
                    else:
                        b = b + 1
                else:
                    b = b + 1

            # print(dateCellValue)

            # print(dateCellValue[3])
            # print(dateCellValue[4])
            if dateCellValue[3] == '0':
                sheet = dateCellValue[4]
            else:
                dateCellValue_0 = dateCellValue[3]
                dateCellValue_1 = dateCellValue[4]
                sheet = ''.join(dateCellValue_0 + dateCellValue_1)



        if sheet == x:

            for index in range(5, 255):
                if type(TypingReport.cell(row=o, column=6).value) is not None:
                    if (TypingReport.cell(row=o, column=6).value) is not None:

                        if (type(TypingReport.cell(row=o, column=6).value)) is str:
                            strOrigin = (TypingReport.cell(row=o, column=6).value)
                        else:
                            strOrigin = (TypingReport.cell(row=o, column=6).value.strftime('%H:%M:%S'))

                        if (type(fileAfterMonitoring2[x].cell(row=index, column=9).value)) is str and ((fileAfterMonitoring2[x].cell(row=index, column=9).value)) is not None:
                            strDestination = fileAfterMonitoring2[x].cell(row=index, column=9).value
                        else:
                            if ((fileAfterMonitoring2[x].cell(row=index, column=9).value)) is not None:
                                strDestination = fileAfterMonitoring2[x].cell(row=index, column=9).value.strftime('%H:%M:%S')


                        if strOrigin == strDestination:
                            fileAfterMonitoring2[x].cell(row=index, column=9).font = copy((TypingReport.cell(row=o, column=6)).font)

                        if (type(TypingReport.cell(row=o, column=7).value)) is str:
                            strOrigin = (TypingReport.cell(row=o, column=7).value)
                        else:
                            strOrigin = (TypingReport.cell(row=o, column=7).value.strftime('%H:%M:%S'))
                        if (type(fileAfterMonitoring2[x].cell(row=index, column=10).value)) is str and (fileAfterMonitoring2[x].cell(row=index, column=10).value) is not None:
                            strDestination = fileAfterMonitoring2[x].cell(row=index, column=10).value
                        else:
                            if (fileAfterMonitoring2[x].cell(row=index, column=10).value) is not None:
                                strDestination = fileAfterMonitoring2[x].cell(row=index, column=10).value.strftime('%H:%M:%S')


                        if type(TypingReport.cell(row=o, column=7).value) is not None:

                            if strOrigin == strDestination:
                                fileAfterMonitoring2[x].cell(row=index, column=10).font = copy((TypingReport.cell(row=o, column=7)).font)

            for index in range(5, 255):
                if type(TypingReport.cell(row=o, column=6).value) is not None:
                    if (TypingReport.cell(row=o, column=6).value) is not None:

                        if (type(TypingReport.cell(row=o, column=6).value)) is str:
                            strOrigin = (TypingReport.cell(row=o, column=6).value)
                        else:
                            strOrigin = (TypingReport.cell(row=o, column=6).value.strftime('%H:%M:%S'))

                        if (type(fileAfterMonitoring2[x].cell(row=index, column=22).value)) is str and (fileAfterMonitoring2[x].cell(row=index, column=22).value) is not None:
                            strDestination = fileAfterMonitoring2[x].cell(row=index, column=22).value
                        else:
                            if (fileAfterMonitoring2[x].cell(row=index, column=22).value) is not None:
                                strDestination = fileAfterMonitoring2[x].cell(row=index, column=22).value.strftime('%H:%M:%S')

                        if strOrigin == strDestination:
                            fileAfterMonitoring2[x].cell(row=index, column=22).font = copy((TypingReport.cell(row=o, column=6)).font)

                    if type(TypingReport.cell(row=o, column=7).value) is not None:

                        if (type(TypingReport.cell(row=o, column=7).value)) is str and (TypingReport.cell(row=o, column=7).value) is not None:
                            strOrigin = (TypingReport.cell(row=o, column=7).value)
                        else:
                            if (TypingReport.cell(row=o, column=7).value) is not None:
                                strOrigin = (TypingReport.cell(row=o, column=7).value.strftime('%H:%M:%S'))

                        if (type(fileAfterMonitoring2[x].cell(row=index, column=23).value)) is str and ((fileAfterMonitoring2[x].cell(row=index, column=23).value)) is not None:
                            strDestination = fileAfterMonitoring2[x].cell(row=index, column=23).value
                        else:
                            if ((fileAfterMonitoring2[x].cell(row=index, column=23).value)) is not None:
                                strDestination = fileAfterMonitoring2[x].cell(row=index, column=23).value.strftime('%H:%M:%S')


                        if strOrigin == strDestination:
                            fileAfterMonitoring2[x].cell(row=index, column=23).font = copy((TypingReport.cell(row=o, column=7).font))



    return

def main(x):
   x = str(x)

   print("I'm into load book")
   loadBook()
   print("I'm into creatingNewWorkBook")
   creatingNewWorkBook()
   print("I'm into variablesPrepraion")
   variablesPrepraion()
   print("I'm into sheetsLoadingTyping south")
   sheetsLoadingTyping(book, 8, exitTimeColumn= 9, arrivalTimeColumn = 10)
   print("I'm into sheetsLoadingfranchisee south")
   sheetsLoadingfranchisee(book, 1, 2, 4, 5, 6, realTimeColumn= 7)
   print("I'm into calculationDiffTime south")
   calculationDiffTime(realTime_1, exitTime_1, realTimeSouth, exitTimeSouth, 11, arrivalTimeNonBlue,  trainNameNonBlue, Date, RB, stopPoint, theoTime, apllTime, south =0)
   print("I'm into sheetsLoadingTyping North")
   sheetsLoadingTyping(book, 21,  exitTimeColumn = 22, arrivalTimeColumn = 23)
   print("I'm into sheetsLoadingfranchisee North")
   sheetsLoadingfranchisee(book, 14, 15, 17, 18, 19, realTimeColumn=20)
   print("I'm into calculationDiffTime north")
   calculationDiffTime(realTime_1, exitTime_1, realTimeNorth, exitTimeNorth, 24, arrivalTimeNonBlue,  trainNameNonBlue, Date, RB, stopPoint, theoTime, apllTime, south = 1)
   paintInOrangeAllTheCellThatWasPainted(exitTimeColumn=9, arrivalTimeColumn=10)
   paintInOrangeAllTheCellThatWasPainted(exitTimeColumn=22, arrivalTimeColumn=23)
   AddingHeadLinesfromFranchiser()
   finalPating()
   return

earliest = input("What is the earliest day of the month from which you would like to run the code?\n")
latest = input("What is the latest day of the month?\n")
fileNameGenerating = input("And last question for now - What is the file name you want to give to the file I'll genearate for you ?\n")

randomNumber = (random.randint(0,3))
if randomNumber == 0:
    print("what a LOVELY name")
elif randomNumber == 1:
    print("This time I will get this name, from here on - please spend a little more effort and originality in choosing the name")
elif randomNumber == 2:
    print("What a b-o-r-i-n-g name !!!!!! :(   ")
else:
    print("ok, this name is ok")
time.sleep(2)



for x in range(int(earliest),int(latest)):
   x= str(x)
   print("I'm in sheet number       ")
   print(x)
   main(x)
   fileAfterMonitoring2.save( fileNameGenerating + '.xlsx')
fileAfterMonitoring2.remove(fileAfterMonitoring2["Sheet"])
fileAfterMonitoring2.save(fileNameGenerating + '.xlsx')

