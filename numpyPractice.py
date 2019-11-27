import openpyxl
import datetime
from copy import copy
import pandas as pd
import numpy as np
from datetime import date, timedelta
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW
import string
import time
import os
originSouth = [1, 2, 3, 7, 8, 9, 10]
destinationSouth = [1, 2, 4, 5, 6, 7, 8]
originNorth = [14, 15, 16, 20, 21, 22, 23]
destinationNorth = [14, 15, 17, 18, 19, 20, 21]

copyFromTyping = [5, 6, 7]
PasteToTypingSouth = [8, 9, 10]
PasteToTypingNorth = [21, 22, 23]


# originSouth = ['A', 'B', 'C', 'G', 'H', 'I']
# destinationSouth = ['A', 'B', 'D', 'E', 'F', 'G']
# originNorth = ['N', 'O', 'P', 'T', 'U', 'V']
# destinationNorth = ['N', 'O', 'Q', 'R', 'S', 'T']


def colorDefinaton():
    global BLACK
    BLACK = 'FF000000'
    global WHITE
    WHITE = 'FFFFFFFF'
    global RED
    RED = 'FFFF0000'
    global DARKRED
    DARKRED = 'FF800000'
    global BLUE
    BLUE = 'FF0000FF'
    global DARKBLUE
    DARKBLUE = 'FF000080'
    global GREEN
    GREEN = 'FF00FF00'
    global DARKGREEN
    DARKGREEN = 'FF008000'
    global YELLOW
    YELLOW = 'FFFFFF00'
    global DARKYELLOW
    DARKYELLOW = 'FF808000'

    return
#C:\Users\shoshana\PycharmProjects\pandas\September2019HeadwayReport.xlsx
def NewFileCreation():
    print("I'm into NewFileCreation")
    global filepath
    filepath = (r"C:\Users\shoshana\PycharmProjects\pandas\SeptemberAfterMonitoring.xlsx")
    global fileAfterMonitoring
    fileAfterMonitoring = openpyxl.Workbook()

    global HeadWay
    partOfFileName = "r" + "'"
    headWayInput =  input("What is the name of the headway report?\n") + ".xlsx"
    HeadWay = openpyxl.load_workbook(headWayInput)
    global HeadWayActive
    HeadWayActive = HeadWay.active
    HeadWayActive
    global sheetsNumber
    sheetsNumber = len(HeadWay.worksheets)

    fileAfterMonitoring.active
    NewFileSheetNumbers = list(range(0, sheetsNumber + 1))



    for sheetIndex in NewFileSheetNumbers:  # go over all sheets
        try:
            i = NewFileSheetNumbers[sheetIndex]
            i = str(i)


        except Exception:
            pass
        fileAfterMonitoring.create_sheet(i)
    # fileAfterMonitoring.remove(fileAfterMonitoring["Sheet"])
    fileAfterMonitoring.remove(fileAfterMonitoring["0"])
    fileAfterMonitoring.remove(fileAfterMonitoring["Sheet"])
    # fileAfterMonitoring.remove(fileAfterMonitoring["111"])

    return

def CopyAndPaste(origin, destination, sheet, fileAfterMonitoring, j):


    for (a) in (range(0, (len(origin)-1))): #go over all columns
        for o in range(1, 250): #go over all cells in a column
            if (not((sheet.cell(row= o, column= origin[a]).value) is None)):
                p = sheet.cell(row=o, column=origin[a]).value
                if ((type(p) is datetime.time) or (type(p) is datetime.datetime)):
                    p = p.strftime('%m/%d/%Y %H:%M:%S')

                    if (("00/01/1900" in p) or ("01/01/1900" in p) or ("01/00/1900" in p) or ("31/12/1899")):
                        d = p[11:19]
                        fileAfterMonitoring[j].cell(row=o, column=destination[a]).value = d
                    else:
                        fileAfterMonitoring[j].cell(row=o, column=destination[a]).value = p
                else:
                    fileAfterMonitoring[j].cell(row=o, column=destination[a]).value = sheet.cell(row=o, column=origin[a]).value


    fileAfterMonitoring.save(filepath)



    return

def TakingDataFromFranchiser(fileAfterMonitoring, filepath):
    print("I'm into TakingDataFromFranchiser")


    for j in range(1, sheetsNumber+1):
        j = str(j)
        global sheet
        sheet = HeadWay[j]
        CopyAndPaste(originSouth, destinationSouth, sheet, fileAfterMonitoring, j)
        CopyAndPaste(originNorth, destinationNorth, sheet, fileAfterMonitoring, j)

    return

def TakingDataFromTypingReport(fileAfterMonitoring, filepath):
    print("I'm into TakingDataFromTypingReport")
    TypingReportLoadinput = input ("what is the name of the typing report file?\n");
    TypingReportLoad = openpyxl.load_workbook(TypingReportLoadinput + ".xlsx", data_only=True)
    TypingReportLoad.active
    global TypingReport
    whatToPaintToBlueFRomTakingDataSouth = 0
    whatToPaintToBlueFRomTakingDataNorth = 0

    global whatToPaintToBlueFRomTakingDataSouthArr
    whatToPaintToBlueFRomTakingDataSouthArr = []

    global whatToPaintToBlueFRomTakingDataNorthArr
    whatToPaintToBlueFRomTakingDataNorthArr = []

    TypingReport = TypingReportLoad["Sheet1"]
    b = 2
    for o in range(2, 10000):

        if (not ((TypingReport.cell(row=o, column=1).value) is None)):
            dateCellValue = TypingReport.cell(row=o, column=1).value
            dateCellValue = dateCellValue.strftime('%m/%d/%Y')
            if (o != 2):
                if o >4:
                    sheetPrevious = sheet
                if ((dateCellValue) != ((TypingReport.cell(row=o - 1, column=1).value).strftime('%m/%d/%Y'))):
                    if ((dateCellValue[3] != '0') and not ((dateCellValue[4] != '1'))):
                        b = 2
                    else:
                        b = b + 1
                else:
                    b = b + 1

            # print(dateCellValue)

            # print(dateCellValue[3])
            # print(dateCellValue[4])
            if dateCellValue[3] == '0':
                sheet = dateCellValue[4]
            else:
                dateCellValue_0 = dateCellValue[3]
                dateCellValue_1 = dateCellValue[4]
                sheet = ''.join(dateCellValue_0 + dateCellValue_1)
            directionCellValue = TypingReport.cell(row=o, column=12).value

            # fileAfterMonitoring[sheet].cell(row =1, column= 8).value = TypingReport.cell(row =1, column= 5).value
            # fileAfterMonitoring[sheet].cell(row =1, column= 8) = copy(TypingReport.cell(row =1, column= 5).font)
            # fileAfterMonitoring[sheet].cell(row =1, column= 8) = copy(TypingReport.cell(row =1, column= 5).border)
            # fileAfterMonitoring[sheet].cell(row =1, column= 8) = copy(TypingReport.cell(row =1, column= 5).fill)
            # fileAfterMonitoring[sheet].cell(row =1, column= 8) = copy(TypingReport.cell(row =1, column= 5).number_format)
            # fileAfterMonitoring[sheet].cell(row =1, column= 8) = copy(TypingReport.cell(row =1, column= 5).protection)
            # fileAfterMonitoring[sheet].cell(row =1, column= 8) = copy(TypingReport.cell(row =1, column= 5).alignment)
########################################################################################################################
########################################################################################################################
########################################################################################################################
            if directionCellValue == "הר הרצל":
                if (TypingReport.cell(row=o - 1, column=12).value == "פסגת זאב"):
                    b = 2

                fileAfterMonitoring[sheet].cell(row=b + 3, column=9).font = copy((TypingReport.cell(row=o, column=6)).font)
                if o > 4:

                    try:
                        if sheet == sheetPrevious:
                            if fileAfterMonitoring[sheet].cell(row=b + 3, column=9).font.color.rgb == "FF00B0F0":
                                whatToPaintToBlueFRomTakingDataSouth = whatToPaintToBlueFRomTakingDataSouth +1
                        else:
                            whatToPaintToBlueFRomTakingDataSouthArr.append(whatToPaintToBlueFRomTakingDataSouth)
                            whatToPaintToBlueFRomTakingDataSouth = 0
                    except AttributeError:
                        whatToPaintToBlueFRomTakingDataSouth = whatToPaintToBlueFRomTakingDataSouth

                fileAfterMonitoring[sheet].cell(row=b + 3, column=9).fill = copy((TypingReport.cell(row=o, column=6)).fill)

                if (type(TypingReport.cell(row=o, column=6).value) is datetime.time) or (type(TypingReport.cell(row=o, column=6).value) is datetime.datetime):
                    a = TypingReport.cell(row=o, column=6).value.strftime('%m/%d/%Y %H:%M:%S')
                    ko = "i did a job"
                else:
                    a = TypingReport.cell(row=o, column=6).value
                    ko = "i didnt do a job"
                if (("00/01/1900" in a) or ("01/01/1900" in a) or ("01/00/1900" in a)):
                    d = a[11:19]

                    fileAfterMonitoring[sheet].cell(row=b + 3, column=9).value = d
                else:
                    fileAfterMonitoring[sheet].cell(row=b + 3, column=9).value = TypingReport.cell(row=o, column=6).value
                ########################################################################################################################
                #if (not(TypingReport.cell(row=o, column=7).value) is None):
                    #print(type(TypingReport.cell(row=o, column=7)).font.color.rgb)
                #if (TypingReport.cell(row=o, column=7).font.color.rgb != "FF000000"):
                    #if (not(type is str)):
                        #v = (TypingReport.cell(row=o, column=7).value.strftime('%H:%M:%S'))
                        #print(v.font.color.rgb)
                    #else:
                        #print(TypingReport.cell(row=o, column=7).font.color.rgb)

                fileAfterMonitoring[sheet].cell(row=b + 3, column=10).font = copy(TypingReport.cell(row=o, column=7).font)
                fileAfterMonitoring[sheet].cell(row=b + 3, column=10).fill = copy(TypingReport.cell(row=o, column=7).fill)
                if (type(TypingReport.cell(row=o, column=7).value)) is datetime.time or (type(TypingReport.cell(row=o, column=7).value) is datetime.datetime):
                    a = TypingReport.cell(row=o, column=7).value.strftime('%m/%d/%Y %H:%M:%S')
                else:
                    a = TypingReport.cell(row=o, column=7).value

                if (("00/01/1900" in a) or ("01/01/1900" in a) or ("01/00/1900" in a)):
                    d = a[11:19]
                    fileAfterMonitoring[sheet].cell(row=b + 3, column=10).value = d

                else:
                    fileAfterMonitoring[sheet].cell(row=b + 3, column=10).value = TypingReport.cell(row=o, column=7).value

########################################################################################################################
                fileAfterMonitoring[sheet].cell(row=b + 3, column=8).font = copy(TypingReport.cell(row=o, column=5).font)
                fileAfterMonitoring[sheet].cell(row=b + 3, column=8).fill = copy(TypingReport.cell(row=o, column=5).fill)
                fileAfterMonitoring[sheet].cell(row=b + 3, column=8).value = TypingReport.cell(row=o, column=5).value
########################################################################################################################
########################################################################################################################
########################################################################################################################
            elif directionCellValue == "פסגת זאב":
                if (TypingReport.cell(row=o - 1, column=12).value == "הר הרצל"):
                    b = 2
                fileAfterMonitoring[sheet].cell(row=b + 3, column=22).font = copy(TypingReport.cell(row=o, column=6).font)
                try:
                    if o > 4:
                        if sheet == sheetPrevious:
                            if fileAfterMonitoring[sheet].cell(row=b + 3, column=9).font.color.rgb == "FF00B0F0":
                                whatToPaintToBlueFRomTakingDataNorth = whatToPaintToBlueFRomTakingDataNorth +1
                        else:
                            whatToPaintToBlueFRomTakingDataNorthArr.append(whatToPaintToBlueFRomTakingDataNorth)
                            whatToPaintToBlueFRomTakingDataNorth = 0
                except AttributeError:
                    whatToPaintToBlueFRomTakingDataNorth = whatToPaintToBlueFRomTakingDataNorth

                fileAfterMonitoring[sheet].cell(row=b + 3, column=22).fill = copy(TypingReport.cell(row=o, column=6).fill)
                if (type(TypingReport.cell(row=o, column=6).value) is datetime.time) or (type(TypingReport.cell(row=o, column=6).value) is datetime.datetime):
                    a = TypingReport.cell(row=o, column=6).value.strftime('%m/%d/%Y %H:%M:%S')
                else:
                    a = TypingReport.cell(row=o, column=6).value

                if (("00/01/1900" in a) or ("01/01/1900" in a) or ("01/00/1900" in a)):
                    d = a[11:19]
                    fileAfterMonitoring[sheet].cell(row=b + 3, column=22).value = d

                else:
                    fileAfterMonitoring[sheet].cell(row=b + 3, column=22).value = TypingReport.cell(row=o, column=6).value

########################################################################################################################
                fileAfterMonitoring[sheet].cell(row=b + 3, column=23).font = copy(TypingReport.cell(row=o, column=7).font)
                fileAfterMonitoring[sheet].cell(row=b + 3, column=23).fill = copy(TypingReport.cell(row=o, column=7).fill)
                if (type(TypingReport.cell(row=o, column=7).value) is datetime.time) or (type(TypingReport.cell(row=o, column=7).value) is datetime.datetime):
                    a = TypingReport.cell(row=o, column=7).value.strftime('%m/%d/%Y %H:%M:%S')
                else:
                    a = TypingReport.cell(row=o, column=7).value
                if (("00/01/1900" in a) or ("01/01/1900" in a) or ("01/00/1900" in a)):
                    d = a[11:19]
                    fileAfterMonitoring[sheet].cell(row=b + 3, column=23).value = d

                else:
                    fileAfterMonitoring[sheet].cell(row=b + 3, column=23).value = TypingReport.cell(row=o, column=7).value

########################################################################################################################
                fileAfterMonitoring[sheet].cell(row=b + 3, column=21).font = copy(TypingReport.cell(row=o, column=5).font)
                fileAfterMonitoring[sheet].cell(row=b + 3, column=21).fill = copy(TypingReport.cell(row=o, column=5).fill)
                fileAfterMonitoring[sheet].cell(row=b + 3, column=21).value = TypingReport.cell(row=o, column=5).value
########################################################################################################################
            else:
                Huston = 1
                print("Huston u have a porblem")
                print(Huston)
                Huston = Huston + 1

    fileAfterMonitoring.save(filepath)
    return
def RealTimeAndExitTimeComperation():

    return

def paintUnusualOriginSationInRed():
    print("I'm into paintUnusualOriginSationInRed")
    # fileAfterMonitoring = openpyxl.Workbook()
    # fileAfterMonitoring.active
    for j in range(1, sheetsNumber+1): #go over all sheets

        for o in range(2, 250):
            j = str(j)
            #HeadWaySheet = HeadWay[j]
            fileAfterMonitoringSheet = fileAfterMonitoring[j]

            if ((not (fileAfterMonitoringSheet.cell( row = o+3, column = 4).value) is None)):
                if ((fileAfterMonitoringSheet.cell( row = o+3, column = 4).value) != "S01KHEI2") and ((fileAfterMonitoringSheet.cell( row = o+3, column = 4).value) != "S01KHEI1"):
                    fileAfterMonitoringSheet.cell( row = o+3, column = 4).font = Font(color=RED)

                elif("S01KHEI" in (fileAfterMonitoringSheet.cell(row=o + 3, column=4).value)):
                    fileAfterMonitoringSheet.cell(row=o + 3, column=4).font = Font(color=BLACK)

                # elif ("KHEI1" in(fileAfterMonitoringSheet.cell(row=o + 3, column=4).value)):
                #     fileAfterMonitoringSheet.cell(row=o + 3, column=4).font = Font(color=BLACK)


            if ((not (fileAfterMonitoringSheet.cell(row=o+3, column=17).value) is None)):
                if ((fileAfterMonitoringSheet.cell(row=o+3, column=17).value) != "S23HERZ1" ):

                    fileAfterMonitoringSheet.cell(row=o+3, column=17).font = Font(color=RED)
                elif ((fileAfterMonitoringSheet.cell(row=o+3, column=17).value) == "S23HERZ1"):
                    fileAfterMonitoringSheet.cell(row=o + 3, column=17).font = Font(color= BLACK)

    fileAfterMonitoring.save(filepath)
    return


def CopyAndPasteForSorting(destination, arr, file, j):
    arr = arr[0]
    for o in range(0, 250):
        try:
            file.cell(row=o + 5, column= destination).value = arr[o]
        except Exception:
            pass

    return


def SortFranchiserDataByColorSouth():
     print("I'm into SortFranchiserDataByColorSouth")
     #fileAfterMonitoring = openpyxl.load_workbook(r"C:\Users\shoshana\PycharmProjects\pandas\julyAfterMonitoring.xlsx")
     for j in range(1, sheetsNumber + 1):  # go over all sheets
         j = str(j)
         fileAfterMonitoringSheet = fileAfterMonitoring[j]
         red = []
         nonRed = []
         dateRed = []
         dateNonRed = []
         sortedDate = []
         sortedColumnByColor = []
         RBRed = []
         RBNonRed = []
         RBSorted = []
         TheoTimeRed = []
         TheoTimeNonRed = []
         TheoTimeSorted = []
         ApllRed = []
         ApllNonRed = []
         ApllSorted = []
         RealTimeRed =[]
         RealTimeNonRed = []
         RealTimeSorted = []

         for o in range(2, 250): # go over all column
             if ((fileAfterMonitoringSheet.cell(row=o+3, column=4).value) is not None):
                  if((fileAfterMonitoringSheet.cell( row = o+3, column = 4).value) != "S01KHEI2") and ((fileAfterMonitoringSheet.cell( row = o+3, column = 4).value) != "S01KHEI1"):

                     red.append(fileAfterMonitoringSheet.cell(row=o+3, column=4).value)
                     dateRed.append(fileAfterMonitoringSheet.cell(row=o+3, column=1).value)
                     RBRed.append(fileAfterMonitoringSheet.cell(row=o+3, column=2).value)
                     TheoTimeRed.append(fileAfterMonitoringSheet.cell(row=o+3, column=5).value)
                     ApllRed.append(fileAfterMonitoringSheet.cell(row=o+3, column=6).value)
                     RealTimeRed.append(fileAfterMonitoringSheet.cell(row=o+3, column=7).value)
                  else:
                     nonRed.append((fileAfterMonitoringSheet.cell(row=o+3, column=4).value))
                     dateNonRed.append((fileAfterMonitoringSheet.cell(row=o+3, column=1).value))
                     RBNonRed.append((fileAfterMonitoringSheet.cell(row=o+3, column=2).value))
                     TheoTimeNonRed.append((fileAfterMonitoringSheet.cell(row=o+3, column=5).value))
                     ApllNonRed.append((fileAfterMonitoringSheet.cell(row=o+3, column=6).value))
                     RealTimeNonRed.append((fileAfterMonitoringSheet.cell(row=o+3, column=7).value))

         if j =='18':
             print(RealTimeNonRed)
             print("////////////////////")
             print(RealTimeRed)
             print("--------------------")
             time.sleep(5)

         sortedColumnByColor = nonRed + red
         sortedDate = dateNonRed + dateRed
         RBSorted = RBNonRed + RBRed
         TheoTimeSorted = TheoTimeNonRed + TheoTimeRed
         ApllSorted = ApllNonRed + ApllRed
         RealTimeSorted = RealTimeNonRed + RealTimeRed

         sortedArrays = np.array([[sortedDate], [RBSorted], [sortedColumnByColor], [TheoTimeSorted], [ApllSorted], [RealTimeSorted]])

         PastingSouthSorted = [1, 2, 4, 5, 6, 7]

         for k in (range(0, 6)):
            CopyAndPasteForSorting(PastingSouthSorted[k], sortedArrays[k], fileAfterMonitoringSheet, j)

     fileAfterMonitoring.save(filepath)
     return


def SortFranchiserDataByColorNorth():
    print("I'm into SortFranchiserDataByColorNorth")
    #fileAfterMonitoring = openpyxl.load_workbook(r"C:\Users\shoshana\PycharmProjects\pandas\julyAfterMonitoring.xlsx")
    for j in range(1, sheetsNumber + 1):  # go over all sheets
        j = str(j)
        fileAfterMonitoringSheet = fileAfterMonitoring[j]
        red = []
        nonRed = []
        dateRed = []
        dateNonRed = []
        sortedDate = []
        sortedColumnByColor = []
        RBRed = []
        RBNonRed = []
        RBSorted = []
        TheoTimeRed = []
        TheoTimeNonRed = []
        TheoTimeSorted = []
        ApllRed = []
        ApllNonRed = []
        ApllSorted = []
        RealTimeRed = []
        RealTimeNonRed = []
        RealTimeSorted = []

        for o in range(2, 250):  # go over all column
            if ((fileAfterMonitoringSheet.cell(row=o + 3, column=17).value) is not None):
                if((fileAfterMonitoringSheet.cell( row = o+3, column = 17).value) != "S01KHEI2") and ((fileAfterMonitoringSheet.cell( row = o+3, column = 17).value) != "S01KHEI1"):

                    red.append(fileAfterMonitoringSheet.cell(row=o + 3, column=17).value)
                    dateRed.append(fileAfterMonitoringSheet.cell(row=o + 3, column=14).value)
                    RBRed.append(fileAfterMonitoringSheet.cell(row=o + 3, column=15).value)
                    TheoTimeRed.append(fileAfterMonitoringSheet.cell(row=o + 3, column=18).value)
                    ApllRed.append(fileAfterMonitoringSheet.cell(row=o + 3, column=19).value)
                    RealTimeRed.append(fileAfterMonitoringSheet.cell(row=o + 3, column=20).value)
                else:
                    nonRed.append((fileAfterMonitoringSheet.cell(row=o + 3, column=17).value))
                    dateNonRed.append((fileAfterMonitoringSheet.cell(row=o + 3, column=14).value))
                    RBNonRed.append((fileAfterMonitoringSheet.cell(row=o + 3, column=15).value))
                    TheoTimeNonRed.append((fileAfterMonitoringSheet.cell(row=o + 3, column=18).value))
                    ApllNonRed.append((fileAfterMonitoringSheet.cell(row=o + 3, column=19).value))
                    RealTimeNonRed.append((fileAfterMonitoringSheet.cell(row=o + 3, column=20).value))

        sortedColumnByColor = nonRed + red
        sortedDate = dateNonRed + dateRed
        RBSorted = RBNonRed + RBRed
        TheoTimeSorted = TheoTimeNonRed + TheoTimeRed
        ApllSorted = ApllNonRed + ApllRed
        RealTimeSorted = RealTimeNonRed + RealTimeRed

        sortedArrays = np.array([[sortedDate], [RBSorted], [sortedColumnByColor], [TheoTimeSorted], [ApllSorted], [RealTimeSorted]])
        PastingSouthSorted = [14, 15, 17, 18, 19, 20]

        for k in (range(0, 6)):
            CopyAndPasteForSorting(PastingSouthSorted[k], sortedArrays[k], fileAfterMonitoringSheet, j)



    fileAfterMonitoring.save(filepath)
    return



def SortAndPaintToBlueSouth():
    for j in range(1, sheetsNumber + 1):  # go over all sheets
        j = str(j)
        counter = 0
        blue = []
        nonblue = []
        combinedBlue = []
        whatToPaintInBlue = []
        trainNameBlue = []
        trainNameNonBlue = []
        trainNameCombined = []
        arraivalTimeBlue = []
        arraivalTimeNonBlue = []
        arraivalTimeCombined = []

        for o in range(1, 250):

            if (fileAfterMonitoring[j].cell(row=o + 4, column=9).value) is not None:

                try:

                    if (fileAfterMonitoring[j].cell(row=o + 4, column=9).font.color.rgb) == "FF00B0F0" or (fileAfterMonitoring[j].cell(row=o + 4, column=9).font.color.rgb) == "FF0070C0":
                        whatToPaintInBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=9).value)
                        counter = counter + 1

                        blue.append(fileAfterMonitoring[j].cell(row=o + 4, column=9).value)
                        trainNameBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=8).value)
                        arraivalTimeBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=10).value)
                    else:
                        nonblue.append(fileAfterMonitoring[j].cell(row=o + 4, column=9).value)
                        trainNameNonBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=8).value)
                        arraivalTimeNonBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=10).value)

                except AttributeError:

                    nonblue.append(fileAfterMonitoring[j].cell(row=o + 4, column=9).value)
                    trainNameNonBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=8).value)
                    arraivalTimeNonBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=10).value)


            else:
                if (fileAfterMonitoring[j].cell(row=o + 4, column=9).value) is not None:
                    nonblue.append(fileAfterMonitoring[j].cell(row=o + 4, column=9).value)
                    trainNameNonBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=8).value)
                    arraivalTimeNonBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=10).value)


        combinedBlue = nonblue + blue
        trainNameCombined = trainNameNonBlue + trainNameBlue
        arraivalTimeCombined = arraivalTimeNonBlue + arraivalTimeBlue

        for o in range(1, 250):
            try:
                (fileAfterMonitoring[j].cell(row=o + 4, column=8).value) = trainNameCombined[o-1]
                (fileAfterMonitoring[j].cell(row=o + 4, column=9).value) = combinedBlue[o-1]
                (fileAfterMonitoring[j].cell(row=o + 4, column=10).value) = arraivalTimeCombined[o-1]
            except Exception:
                pass
        j = int(j)
        if j -1 < 28:

            for i in range(0, whatToPaintToBlueFRomTakingDataSouthArr[j-1]):

                j = str(j)
                (fileAfterMonitoring[j].cell(row=((len(combinedBlue)) + 5 - i), column=9)).font = Font(color="FF00B0F0")


        if type(j) is not str:
            j = str(j)
        for i in range(0, len(nonblue)):
            (fileAfterMonitoring[j].cell(row=( i + 5), column=9)).font = Font(color="FF000000")



    return


def SortAndPaintToBlueNorth():
    for j in range(1, sheetsNumber + 1):  # go over all sheets
        j = str(j)
        counter = 0
        blue = []
        nonblue = []
        combinedBlue = []
        whatToPaintInBlue = []
        trainNameBlue = []
        trainNameNonBlue = []
        trainNameCombined = []
        arraivalTimeBlue = []
        arraivalTimeNonBlue = []
        arraivalTimeCombined = []

        for o in range(1, 250):

            if (fileAfterMonitoring[j].cell(row=o + 4, column=22).value) is not None:
                paintedInBlue = (fileAfterMonitoring[j].cell(row=o + 4, column=22).value)

                try:
                    if (fileAfterMonitoring[j].cell(row=o + 4, column=22).font.color.rgb) == "FF00B0F0" or (fileAfterMonitoring[j].cell(row=o + 4, column=22).font.color.rgb) == "FF0070C0":

                        whatToPaintInBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=22).value)
                        counter = counter + 1
                        blue.append(fileAfterMonitoring[j].cell(row=o + 4, column=22).value)
                        trainNameBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=21).value)
                        arraivalTimeBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=23).value)
                    else:
                        nonblue.append(fileAfterMonitoring[j].cell(row=o + 4, column=22).value)
                        trainNameNonBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=21).value)
                        arraivalTimeNonBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=23).value)
                except AttributeError:
                    nonblue.append(fileAfterMonitoring[j].cell(row=o + 4, column=22).value)
                    trainNameNonBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=21).value)
                    arraivalTimeNonBlue.append(fileAfterMonitoring[j].cell(row=o + 4, column=23).value)

        combinedBlue = nonblue + blue
        trainNameCombined = trainNameNonBlue + trainNameBlue
        arraivalTimeCombined = arraivalTimeNonBlue + arraivalTimeBlue


        for o in range(1, 250):
            try:
                (fileAfterMonitoring[j].cell(row=o + 4, column=21).value) = trainNameCombined[o-1]
                (fileAfterMonitoring[j].cell(row=o + 4, column=22).value) = combinedBlue[o-1]
                (fileAfterMonitoring[j].cell(row=o + 4, column=23).value) = arraivalTimeCombined[o-1]
            except Exception:
                pass
        j = int(j)
        if 0<j-1 <28:

            try:
                for i in range(0, whatToPaintToBlueFRomTakingDataNorthArr[j-1]):
                    j = str(j)
                    (fileAfterMonitoring[j].cell(row=((len(combinedBlue)) + 5 - i), column=22)).font = Font(color="FF00B0F0")

            except IndexError:
                for i in range(0, len(whatToPaintInBlue)):
                    if j is not str:
                        j = str(j)
                    (fileAfterMonitoring[j].cell(row=((len(combinedBlue)) + 5 - i), column=22)).font = Font(color="FF00B0F0")

        if j is not str:
            j = str(j)
        for i in range(0, len(nonblue)):
            (fileAfterMonitoring[j].cell(row=( i + 5), column=22)).font = Font(color="FF000000")



    return



def AddingHeadLinesfromTypingReport():
    fromWhereToTake = [5, 6, 7, 5, 6, 7]
    whereToPaste = [8, 9, 10, 21, 22, 23]
    for j in range(1, 32):
        j = str(j)

        for i in range(0, 6):
            if j != '32':

                fileAfterMonitoring[j].cell(row=4, column=8).value = TypingReport.cell(row=1, column=5).value
                fileAfterMonitoring[j].cell(row=4, column=9).value = TypingReport.cell(row=1, column=6).value
                fileAfterMonitoring[j].cell(row=4, column=10).value = TypingReport.cell(row=1, column=7).value

                fileAfterMonitoring[j].cell(row=4, column=21).value = TypingReport.cell(row=1, column=5).value
                fileAfterMonitoring[j].cell(row=4, column=22).value = TypingReport.cell(row=1, column=6).value
                fileAfterMonitoring[j].cell(row=4, column=23).value = TypingReport.cell(row=1, column=7).value

                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).font = copy(TypingReport.cell(row=1, column= fromWhereToTake[i]).font)
                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).border = copy(TypingReport.cell(row=1, column= fromWhereToTake[i]).border)
                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).fill = copy(TypingReport.cell(row=1, column= fromWhereToTake[i]).fill)
                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).number_format = copy(TypingReport.cell(row=1, column= fromWhereToTake[i]).number_format)
                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).protection = copy(TypingReport.cell(row=1, column= fromWhereToTake[i]).protection)
                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).alignment = copy(TypingReport.cell(row=1, column= fromWhereToTake[i]).alignment)

    return

def AddingHeadLinesfromFranchiser():
    fromWhereToTake = [1, 2, 3 , 7, 8, 9, 1, 2, 3, 7, 8, 9]
    whereToPaste = [1, 2, 4, 5, 6, 7, 14, 15, 17, 18, 19, 20]
    #style = [font, border, fill, number_format, protection, alignment]
    for j in range(1, 32):
        if j != 32:
            j = str(j)


            for i in range(0, 12):
                fileAfterMonitoring[j].cell(row=4, column=1).value = HeadWay[j].cell(row=4, column=1).value
                fileAfterMonitoring[j].cell(row=4, column=2).value = HeadWay[j].cell(row=4, column=2).value
                fileAfterMonitoring[j].cell(row=4, column=4).value = HeadWay[j].cell(row=4, column=3).value
                fileAfterMonitoring[j].cell(row=4, column=5).value = HeadWay[j].cell(row=4, column=7).value
                fileAfterMonitoring[j].cell(row=4, column=6).value = HeadWay[j].cell(row=4, column=8).value
                fileAfterMonitoring[j].cell(row=4, column=7).value = HeadWay[j].cell(row=4, column=9).value

                fileAfterMonitoring[j].cell(row=4, column=14).value = HeadWay[j].cell(row=4, column=1).value
                fileAfterMonitoring[j].cell(row=4, column=15).value = HeadWay[j].cell(row=4, column=2).value
                fileAfterMonitoring[j].cell(row=4, column=17).value = HeadWay[j].cell(row=4, column=3).value
                fileAfterMonitoring[j].cell(row=4, column=18).value = HeadWay[j].cell(row=4, column=7).value
                fileAfterMonitoring[j].cell(row=4, column=19).value = HeadWay[j].cell(row=4, column=8).value
                fileAfterMonitoring[j].cell(row=4, column=20).value = HeadWay[j].cell(row=4, column=9).value

                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).font = copy(HeadWay[j].cell(row=4, column= fromWhereToTake[i]).font)
                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).border = copy(HeadWay[j].cell(row=4, column= fromWhereToTake[i]).border)
                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).fill = copy(HeadWay[j].cell(row=4, column= fromWhereToTake[i]).fill)
                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).number_format = copy(HeadWay[j].cell(row=4, column= fromWhereToTake[i]).number_format)
                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).protection = copy(HeadWay[j].cell(row=4, column= fromWhereToTake[i]).protection)
                fileAfterMonitoring[j].cell(row=4, column= whereToPaste[i]).alignment = copy(HeadWay[j].cell(row=4, column= fromWhereToTake[i]).alignment)

    return

def main():
    colorDefinaton()
    NewFileCreation()
    TakingDataFromFranchiser(fileAfterMonitoring, filepath)
    TakingDataFromTypingReport(fileAfterMonitoring, filepath)
    #RealTimeAndExitTimeComperation()
    paintUnusualOriginSationInRed()
    SortFranchiserDataByColorSouth()
    SortFranchiserDataByColorNorth()
    paintUnusualOriginSationInRed()
    SortAndPaintToBlueSouth()
    SortAndPaintToBlueNorth()
    AddingHeadLinesfromTypingReport()
    #fileAfterMonitoring.save(filepath)
    AddingHeadLinesfromFranchiser()
    fileAfterMonitoring.save(r"C:\Users\shoshana\PycharmProjects\pandas\SeptemberAfterMonitoringBeforeManipilation.xlsx")
    os.system('python diffCalculation.py')
    return

main()

